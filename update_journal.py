#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
プログラム②：日誌更新

役割
- 授業日誌Excelの「日付・曜日」「講師名」だけを最新スケジュールに合わせて更新する
- 先生が記入した本文は絶対に消さない

保護対象（触らない）
- D7   授業内容
- F8   ページ
- E9:E10 宿題
- E11  記録
- D17  録画URL

更新対象（ブロック先頭基準）
- 月   : top+4, left   （例: B10）
- 日   : top+5, left   （例: B11）
- 曜日 : top+6, left   （例: B12）
- 講師 : top+8, left   （例: B14）

※ 先生用テンプレの見た目上は「C9-11 / B15」と認識されやすいですが、
   実ファイル上は export_by_grade_subject.py がこの座標に書き込んでいるため、
   互換性維持のため同じ座標を使います。

前提
- 同じフォルダに export_by_grade_subject.py があること
- 授業日誌Excelは既に生成済みであること
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# export_by_grade_subject.py から、既に動いている読み取りロジックをそのまま使う
from export_by_grade_subject import (  # type: ignore
    GRADES,
    SUBJS,
    GRADE_LABEL,
    _display_subj_name,
    choose_target_sheets,
    collect_events,
    pick_schedule_in_same_folder,
)

ROW_STEP = 20
FIRST_BLOCK_COL = 2
BLOCK_WIDTH = 10
BASE_TOP_MAIN = 6
BASE_TOP_X = 6
CLASS_ORDER = ["S", "A", "B"]
CLASS_INDEX = {k: i for i, k in enumerate(CLASS_ORDER)}

MONTH_ROW_OFFSET = 4
DAY_ROW_OFFSET = 5
WDAY_ROW_OFFSET = 6
TEACHER_ROW_OFFSET = 8

TARGET_FOLDER_NAME = "09　授業日誌"
MONTH_SHEET_RE = re.compile(r"^(\d{4})-(\d{2})(?:\(再(\d+)\))?$")


def get_default_journal_dir() -> Optional[Path]:
    user_home = Path.home()
    candidates = [
        user_home / "OneDrive" / "●勉強クラブ共有" / TARGET_FOLDER_NAME,
        Path(r"C:\Users\kudok\OneDrive\●勉強クラブ共有\09　授業日誌"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def merged_top_left_cell(ws, row: int, col: int):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col)
    return ws.cell(row, col)


def safe_write(ws, row: int, col: int, value) -> None:
    merged_top_left_cell(ws, row, col).value = value


def clear_update_fields(ws, top_row: int, left_col: int) -> None:
    safe_write(ws, top_row + MONTH_ROW_OFFSET, left_col, "")
    safe_write(ws, top_row + DAY_ROW_OFFSET, left_col, "")
    safe_write(ws, top_row + WDAY_ROW_OFFSET, left_col, "")
    safe_write(ws, top_row + TEACHER_ROW_OFFSET, left_col, "")


def write_update_fields(ws, top_row: int, left_col: int, month: int, day, wday: str, teacher: str) -> None:
    safe_write(ws, top_row + MONTH_ROW_OFFSET, left_col, str(month) if month else "")
    safe_write(ws, top_row + DAY_ROW_OFFSET, left_col, str(day) if day not in (None, "") else "")
    safe_write(ws, top_row + WDAY_ROW_OFFSET, left_col, wday or "")
    safe_write(ws, top_row + TEACHER_ROW_OFFSET, left_col, teacher or "")


def find_month_sheet(wb: openpyxl.Workbook, year: int, month: int):
    target = f"{year:04d}-{month:02d}"
    if target in wb.sheetnames:
        return wb[target]

    cands: List[Tuple[int, str]] = []
    for name in wb.sheetnames:
        m = MONTH_SHEET_RE.match(name)
        if not m:
            continue
        y = int(m.group(1))
        mo = int(m.group(2))
        retry = int(m.group(3) or 0)
        if y == year and mo == month:
            cands.append((retry, name))

    if cands:
        cands.sort()
        return wb[cands[-1][1]]
    return None


def count_slots_in_sheet(ws) -> int:
    maxcol = ws.max_column
    return max(1, min(30, (maxcol - FIRST_BLOCK_COL) // BLOCK_WIDTH + 1))


def open_wb(path: Path) -> Optional[openpyxl.Workbook]:
    try:
        return openpyxl.load_workbook(path)
    except Exception as e:
        print(f"[WARN] 開けませんでした: {path.name} / {e}")
        return None


def build_buckets(events) -> Dict[Tuple[str, str, str], list]:
    buckets: Dict[Tuple[str, str, str], list] = {}
    from collections import defaultdict
    dd = defaultdict(list)
    for e in events:
        dd[(e.grade, e.subj, e.klass)].append(e)
    return dd


def find_workbook(journal_dir: Path, filename: str) -> Optional[Path]:
    p = journal_dir / filename
    if p.exists():
        return p
    for candidate in journal_dir.rglob(filename):
        if candidate.is_file():
            return candidate
    return None


def update_main_workbook(path: Path, year: int, month: int, s_list: list, a_list: list, b_list: list) -> bool:
    wb = open_wb(path)
    if wb is None:
        return False

    ws = find_month_sheet(wb, year, month)
    if ws is None:
        print(f"[WARN] 月シートが見つかりません: {path.name} / {year:04d}-{month:02d}")
        wb.close()
        return False

    total_slots = count_slots_in_sheet(ws)
    lens = {"S": len(s_list), "A": len(a_list), "B": len(b_list)}
    classes = {"S": s_list, "A": a_list, "B": b_list}

    for slot in range(total_slots):
        col_left = FIRST_BLOCK_COL + BLOCK_WIDTH * slot

        # まず更新対象だけ全消し（本文は消さない）
        for klass in CLASS_ORDER:
            top = BASE_TOP_MAIN + ROW_STEP * CLASS_INDEX[klass]
            clear_update_fields(ws, top, col_left)

        slot_events = {}
        has_special = False
        for klass in CLASS_ORDER:
            if slot < lens[klass]:
                ev = classes[klass][slot]
                slot_events[klass] = ev
                if getattr(ev, "special", False):
                    has_special = True

        for klass in CLASS_ORDER:
            if slot >= lens[klass]:
                continue
            ev = slot_events[klass]
            if has_special and not getattr(ev, "special", False):
                continue
            top = BASE_TOP_MAIN + ROW_STEP * CLASS_INDEX[klass]
            write_update_fields(ws, top, col_left, month, ev.day, ev.wday, ev.teacher or "")

    wb.save(path)
    wb.close()
    print(f"[OK] 更新: {path.name} / {ws.title}")
    return True


def update_x_workbook(path: Path, year: int, month: int, x_list: list) -> bool:
    wb = open_wb(path)
    if wb is None:
        return False

    ws = find_month_sheet(wb, year, month)
    if ws is None:
        print(f"[WARN] 月シートが見つかりません: {path.name} / {year:04d}-{month:02d}")
        wb.close()
        return False

    total_slots = count_slots_in_sheet(ws)
    for slot in range(total_slots):
        col_left = FIRST_BLOCK_COL + BLOCK_WIDTH * slot
        clear_update_fields(ws, BASE_TOP_X, col_left)
        if slot < len(x_list):
            ev = x_list[slot]
            write_update_fields(ws, BASE_TOP_X, col_left, month, ev.day, ev.wday, ev.teacher or "")

    wb.save(path)
    wb.close()
    print(f"[OK] 更新: {path.name} / {ws.title}")
    return True


def update_hojyu_workbook(path: Path, year: int, month: int, hits: list) -> bool:
    # 補講ファイルは S ブロックだけを使って並べている
    wb = open_wb(path)
    if wb is None:
        return False

    ws = find_month_sheet(wb, year, month)
    if ws is None:
        print(f"[WARN] 月シートが見つかりません: {path.name} / {year:04d}-{month:02d}")
        wb.close()
        return False

    total_slots = count_slots_in_sheet(ws)
    for slot in range(total_slots):
        col_left = FIRST_BLOCK_COL + BLOCK_WIDTH * slot
        clear_update_fields(ws, BASE_TOP_MAIN, col_left)
        if slot < len(hits):
            ev = hits[slot]
            write_update_fields(ws, BASE_TOP_MAIN, col_left, month, ev.day, ev.wday, ev.teacher or "")

    wb.save(path)
    wb.close()
    print(f"[OK] 更新: {path.name} / {ws.title}")
    return True


def process_campus(journal_dir: Path, campus: str, events: list, year: int, month: int) -> None:
    buckets = build_buckets([e for e in events if ("英語補講" not in e.text and "数学補講" not in e.text)])

    for grade in GRADES:
        for subj in SUBJS:
            s_list = buckets.get((grade, subj, "S"), [])
            a_list = buckets.get((grade, subj, "A"), [])
            b_list = buckets.get((grade, subj, "B"), [])
            x_list = buckets.get((grade, subj, "X"), [])

            if not any([s_list, a_list, b_list, x_list]):
                continue

            grade_j = GRADE_LABEL[grade]
            subj_j = _display_subj_name(grade, subj)

            if any([s_list, a_list, b_list]):
                filename = f"{campus}{grade_j}{subj_j}_{year}.xlsx"
                path = find_workbook(journal_dir, filename)
                if path is None:
                    print(f"[WARN] 日誌ファイルが見つかりません: {filename}")
                else:
                    update_main_workbook(path, year, month, s_list, a_list, b_list)

            if x_list:
                filename = f"{campus}{grade_j}X{subj_j}_{year}.xlsx"
                path = find_workbook(journal_dir, filename)
                if path is None:
                    print(f"[WARN] 日誌ファイルが見つかりません: {filename}")
                else:
                    update_x_workbook(path, year, month, x_list)

    for keyword, title in [("英語補講", "英語補講"), ("数学補講", "数学補講")]:
        hits = [e for e in events if keyword in e.text]
        if not hits:
            continue
        filename = f"{campus}{title}_{year}.xlsx"
        path = find_workbook(journal_dir, filename)
        if path is None:
            print(f"[WARN] 日誌ファイルが見つかりません: {filename}")
        else:
            update_hojyu_workbook(path, year, month, hits)


def resolve_schedule_path(arg_schedule: Optional[str], prefer: Optional[str]) -> Tuple[int, int, Path]:
    if arg_schedule:
        p = Path(arg_schedule)
        if not p.exists():
            raise FileNotFoundError(f"スケジュールが見つかりません: {p}")
        m = re.match(r"^(\d{4})年0?(\d{1,2})月スケジュール\.xlsm$", p.name)
        if not m:
            raise ValueError("スケジュール名は 'YYYY年M月スケジュール.xlsm' 形式にしてください。")
        return int(m.group(1)), int(m.group(2)), p

    prefer_months = [int(x) for x in re.split(r"[,\s]+", prefer.strip()) if x] if prefer else None
    return pick_schedule_in_same_folder(prefer_months)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--journal-dir", help="授業日誌フォルダ。未指定時は OneDrive 共有フォルダを自動探索")
    ap.add_argument("--schedule", help="対象スケジュール.xlsm のフルパス。未指定時は同フォルダから自動検出")
    ap.add_argument("--prefer", help="優先月（例: 3 または 3,2）")
    args = ap.parse_args()

    journal_dir = Path(args.journal_dir) if args.journal_dir else get_default_journal_dir()
    if journal_dir is None or not journal_dir.exists():
        raise FileNotFoundError("授業日誌フォルダが見つかりません。--journal-dir で指定してください。")

    year, month, schedule_path = resolve_schedule_path(args.schedule, args.prefer)
    print(f"[INFO] 授業日誌フォルダ: {journal_dir}")
    print(f"[INFO] スケジュール: {schedule_path} ({year}-{month:02d})")

    wb_s = openpyxl.load_workbook(schedule_path, data_only=True, keep_vba=True)
    targets = choose_target_sheets(wb_s)
    if not targets:
        print("[WARN] 本校/南教室の教務部用シートが見つかりません。")
        return

    for campus, sname, sh in targets:
        print(f"[INFO] 対象シート: {campus} / {sname}")
        events = collect_events(sh, month)
        process_campus(journal_dir, campus, events, year, month)

    print("[INFO] 完了しました。本文セルは変更していません。")


if __name__ == "__main__":
    main()
