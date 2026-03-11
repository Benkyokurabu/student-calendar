# -*- coding: utf-8 -*-
"""
授業日誌自動生成（レガシー準拠）
- A: 凡例は H56:AF56 を最優先で読む。足りなければ 54〜58行, C〜AZ で補完。
- B: 先生色は「塗りつぶし」だけを使用（フォント色は使わない）。
     一致優先順位：RGB 完全一致 → THEME/INDEXED 完全一致 → RGB近傍（閾値<=24, Euclid）
     その後に 列内継承 → 上方向探索（最大20行）。
- C: 日付＝[左端列, 右から2列目] / 曜日＝[左から2列目, 右端] の固定ルール。
- 読み取るシートは「本校教務部用」「南教室教務部用」に相当する2枚のみ（表記ゆれ対応）。
  （シート名に「教務」を含み、かつ「本校」または「南/南教室」を含むものを採用）
- テンプレ分離：S/A/B= TEMPLATE_MAIN.xlsx、X= TEMPLATE_X.xlsx。出力は各1シート（「n月」）に統一。
- 未使用スロットはクラス欄も含めて空白。録画(B17/C17)・進捗(B20〜G20)・担当セルは保護。
- 結合セルは左上に書く safe_write を使用。空欄は "" を書いて “0” 表示を回避。

※ 追加仕様：
- 小学生（4,5,6）の「数」は、出力ファイル名で「算数」にする（内部処理は従来どおり「数」）
- 出力ファイル名に「_{year}_{month:02d}」を付与する
- 「特」付き授業（Event.special=True）のスロットでは、そのクラスのラベルを「特」にする
"""

import sys, re, argparse
from pathlib import Path
from collections import defaultdict, namedtuple
from typing import Optional, Dict, Tuple, Set, List, Any
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles.colors import COLOR_INDEX
import math

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ===== パス設定（同じフォルダに置く） =====
TEMPLATE_MAIN = Path(r"TEMPLATE_MAIN.xlsx")  # S/A/B 用
TEMPLATE_X    = Path(r"TEMPLATE_X.xlsx")     # X 用

# ===== 定数 =====
ROW_STEP = 20  # 段の高さ（S=6, A=26, B=46, X=66）
GRADES = ["1","2","3","4","5","6"]
SUBJS  = ["英","数","国","理","社"]
GRADE_LABEL = {"1":"中１","2":"中２","3":"中３","4":"小４","5":"小５","6":"小６"}
SUBJ_NAME   = {"英":"英語","数":"数学","国":"国語","理":"理科","社":"社会"}

# RGB 近傍一致の閾値（Euclid 距離）
RGB_NEAR_THRESH = 24

Event = namedtuple("Event", "month day wday time classroom grade klass subj text teacher special r c")

# ===== スケジュール自動選択 =====
RE_SCHEDULE = re.compile(r"^(\d{4})年0?(\d{1,2})月スケジュール\.xlsm$", re.ASCII)
def pick_schedule_in_same_folder(prefer=None):
    here=Path("."); cands=[]
    for p in here.iterdir():
        if p.is_file():
            m=RE_SCHEDULE.match(p.name)
            if m: cands.append((int(m.group(1)), int(m.group(2)), p))
    if not cands:
        raise FileNotFoundError("同フォルダに『YYYY年M月スケジュール.xlsm』が見つかりません。")
    if prefer:
        for pm in prefer:
            for (y,mo,pp) in sorted(cands,key=lambda x:(x[0],x[1]), reverse=True):
                if mo==pm: return y,mo,pp
    y,mo,pp = sorted(cands,key=lambda x:(x[0],x[1]), reverse=True)[0]
    return y,mo,pp

def month_sheet_name(m:int)->str: return f"{m}月"

# ===== ユーティリティ =====
def cell_text(c):
    v=c.value
    return "" if v is None else str(v).strip()

def normalize_digits(s):
    return s.translate(str.maketrans({chr(ord('０')+i):str(i) for i in range(10)}))

def parse_class_token(txt):
    s=normalize_digits(txt).replace(" ","").replace("\u3000","")
    if any(x in s for x in ("休講","休み","休校","休","テスト","模試")):
        return None
    m=re.search(r"([1-6])\s*([SABXＳＡＢＸ])?(特)?\s*([数算国英理社])", s)
    if not m: return None
    grade=m.group(1)
    klass=(m.group(2) or "").upper().replace("Ｓ","S").replace("Ａ","A").replace("Ｂ","B").replace("Ｘ","X")
    special=bool(m.group(3))
    subj=m.group(4)
    subj_norm="数" if subj in ("数","算") else subj
    return grade,klass,subj_norm,special

def safe_write(ws,row,col,val):
    # 結合セルは左上へ書く
    for rng in ws.merged_cells.ranges:
        if rng.min_row<=row<=rng.max_row and rng.min_col<=col<=rng.max_col:
            ws.cell(rng.min_row, rng.min_col).value = val
            return
    ws.cell(row=row, column=col).value = val

# ===== 色キー（塗りのみ：rgb / indexed / theme(+tint)） =====
def _rgb_from_argb(a: Optional[str]) -> Optional[str]:
    if not a: return None
    h=a.upper()
    if len(h)==8: h=h[2:]  # 先頭AAを落とす
    return h if len(h)==6 else None

def _fill_color_keys(fill) -> Set[Tuple[str, Any]]:
    keys=set()
    for col in (getattr(fill,"start_color",None), getattr(fill,"fgColor",None)):
        if not col: continue
        t=getattr(col,"type",None)
        if t=="rgb" and getattr(col,"rgb",None):
            rgb=_rgb_from_argb(col.rgb)
            if rgb: keys.add(("rgb", rgb))
        elif t=="indexed" and getattr(col,"indexed",None) is not None:
            keys.add(("indexed", col.indexed))
        elif t=="theme" and getattr(col,"theme",None) is not None:
            tint = getattr(col,"tint",0.0)
            try: tint_val=float(tint)
            except Exception: tint_val=0.0
            keys.add(("theme", col.theme, round(tint_val,3)))
    return keys

def _hex_to_rgb_tuple(h: str) -> Tuple[int,int,int]:
    return int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)

def _rgb_dist(h1: str, h2: str) -> float:
    r1,g1,b1=_hex_to_rgb_tuple(h1); r2,g2,b2=_hex_to_rgb_tuple(h2)
    return math.sqrt((r1-r2)**2 + (g1-g2)**2 + (b1-b2)**2)

def looks_like_name(s:str)->bool:
    return bool(re.fullmatch(r"[ぁ-んァ-ン一-龥々〆ヵヶA-Za-z]{1,6}", s or ""))

# ===== 先生凡例 =====
def detect_teacher_legend(ws) -> Dict[Tuple[str,Any], str]:
    mapping: Dict[Tuple[str,Any], str] = {}

    H, AF = column_index_from_string("H"), column_index_from_string("AF")
    C, AZ = column_index_from_string("C"), column_index_from_string("AZ")

    # 最優先：H56:AF56
    r = 56
    for c in range(H, AF+1):
        name = cell_text(ws.cell(r,c))
        if not looks_like_name(name): continue
        for k in _fill_color_keys(ws.cell(r,c).fill):
            mapping.setdefault(k, name)

    # 補完：54〜58 行 × C〜AZ 列
    if not mapping:
        for rr in range(54, 59):
            for cc in range(C, AZ+1):
                name = cell_text(ws.cell(rr,cc))
                if not looks_like_name(name): continue
                for k in _fill_color_keys(ws.cell(rr,cc).fill):
                    mapping.setdefault(k, name)

    return mapping

# ===== 先生確定 =====
def resolve_teacher(ws, legend:Dict[Tuple[str,Any],str], row:int, col:int, col_confirmed:Dict[int,str]) -> str:
    cell = ws.cell(row,col)
    keys = _fill_color_keys(cell.fill)

    # 1) 完全一致（RGB優先→THEME/INDEXED）
    for k in keys:
        if k[0]=="rgb" and k in legend:
            name=legend[k]; col_confirmed[col]=name; return name
    for k in keys:
        if k[0] in ("theme","indexed") and k in legend:
            name=legend[k]; col_confirmed[col]=name; return name

    # 2) RGB近傍
    cell_rgbs = [k[1] for k in keys if k[0]=="rgb"]
    if cell_rgbs:
        legend_rgbs = [(k[1], v) for (k,v) in legend.items() if k[0]=="rgb"]
        best = (1e9, "")
        for crgb in cell_rgbs:
            for lrgb, lname in legend_rgbs:
                d = _rgb_dist(crgb, lrgb)
                if d < best[0]:
                    best = (d, lname)
        if best[0] <= RGB_NEAR_THRESH:
            col_confirmed[col] = best[1]
            return best[1]

    # 3) 列継承
    if col in col_confirmed:
        return col_confirmed[col]

    # 4) 上方向探索
    for r in range(row-1, max(1,row-20), -1):
        keys2 = _fill_color_keys(ws.cell(r,col).fill)
        for k in keys2:
            if k[0]=="rgb" and k in legend:
                name=legend[k]; col_confirmed[col]=name; return name
        for k in keys2:
            if k[0] in ("theme","indexed") and k in legend:
                name=legend[k]; col_confirmed[col]=name; return name
        cell_rgbs2 = [k[1] for k in keys2 if k[0]=="rgb"]
        if cell_rgbs2:
            legend_rgbs = [(k[1], v) for (k,v) in legend.items() if k[0]=="rgb"]
            best=(1e9,"")
            for crgb in cell_rgbs2:
                for lrgb,lname in legend_rgbs:
                    d=_rgb_dist(crgb,lrgb)
                    if d<best[0]:
                        best=(d,lname)
            if best[0] <= RGB_NEAR_THRESH:
                col_confirmed[col]=best[1]
                return best[1]

    return ""

# ===== 教室行・時間帯・日付/曜日 =====
def detect_classroom_row(sh):
    C,AD = column_index_from_string("C"), column_index_from_string("AD")
    for r in range(1,60):
        vals=[cell_text(sh.cell(r,c)) for c in range(C,AD+1)]
        if sum(1 for v in vals if v and v[0] in "①②③④⑤⑥⑦⑧⑨")>=3:
            return r
    return None

def find_timeband_above(sh,row,col):
    for mr in sh.merged_cells.ranges:
        if mr.min_col<=col<=mr.max_col and mr.min_row<row:
            txt=cell_text(sh.cell(mr.min_row,mr.min_col))
            m=re.search(r"\d{1,2}[:：]\d{2}\s*[~～]\s*\d{1,2}[:：]\d{2}", txt or "")
            if m: return m.group(0).replace("：",":")
    for r in range(row-1,1,-1):
        v=cell_text(sh.cell(r,col))
        if re.search(r"\d{1,2}[:：]\d{2}\s*[~～]\s*\d{1,2}[:：]\d{2}", v or ""):
            return v.replace("：",":")
    return ""

def _merged_cell_top_left_value(sh, row:int, col:int):
    """結合セルの場合、その結合範囲の左上セルの value を返す。結合でなければ当該セルの value。"""
    for mr in sh.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return sh.cell(mr.min_row, mr.min_col).value
    return sh.cell(row, col).value

def cell_text_with_merge(sh, row:int, col:int) -> str:
    v = _merged_cell_top_left_value(sh, row, col)
    return "" if v is None else str(v).strip()

def parse_day_week_from_row(sh, row, *, allow_prev_row:bool=False):
    """C: 日付=[左端,右から2列目]／曜日=[左から2列目,右端]

    縦結合（2行固定）で下段に日付/曜日が直接入らないケースに対応するため、
    結合セルの左上値を参照して“見た目どおり”に読み取れるようにする。
    allow_prev_row=True のとき、当該行で取れない場合に直上1行も試す（下段救済用）。
    """
    def _read_on_row(rr:int):
        day = None; w = None
        maxcol = sh.max_column
        left_date_col = 1
        right_date_col = max(1, maxcol - 1)
        left_wday_col = 2
        right_wday_col = maxcol

        # 日付
        for c in (left_date_col, right_date_col):
            vv = normalize_digits(cell_text_with_merge(sh, rr, c))
            m = re.search(r"([0-9]{1,2})", vv)
            if m:
                d = int(m.group(1))
                if 1 <= d <= 31:
                    day = d
                    break

        # 曜日
        for c in (left_wday_col, right_wday_col):
            vv = cell_text_with_merge(sh, rr, c)
            for ch in vv:
                if ch in "月火水木金土日":
                    w = ch
                    break
            if w:
                break

        return day, w

    day, w = _read_on_row(row)

    if allow_prev_row and (day is None or w is None) and row > 1:
        d2, w2 = _read_on_row(row - 1)
        if day is None:
            day = d2
        if w is None:
            w = w2

    return day, w

# ===== クリア範囲（録画・進捗・担当を保護） =====
PROTECT_REL_CELLS = {(11,0),(11,1),(8,0)}  # (B17, C17, 担当セル top+8,left)
PROTECT_REL_ROW20_RANGE = (14, 0, 5)       # B20〜G20

CLEAR_RANGES_REL: List[Tuple[int,int,int,int]] = [
    (1,  2,  8, 1),   # D7:J7
    (2,  4,  8, 1),   # F8:J8
    (3,  3,  8, 1),   # E9:J9
    (4,  3,  8, 1),   # E10:J10
    (5,  3,  8, 6),   # E11:J16
    (11, 1,  8, 1),   # C17:J17（C17は個別スキップ）
    (12, 1,  8, 1),   # C18:J18
    (13, 6,  6, 1),   # H19
    (13, 7,  8, 1),   # I19:J19
    (14, 1,  8, 4),   # C20:J23（B20〜G20は個別スキップ）
]

def _is_protected(top:int, left:int, rr:int, cc:int) -> bool:
    if (rr - top, cc - left) in PROTECT_REL_CELLS:
        return True
    pr, pc0, pcw = PROTECT_REL_ROW20_RANGE
    if rr == top + pr and left + pc0 <= cc <= left + pcw:
        return True
    return False

def clear_one_block(ws, top:int, left:int):
    for drow, dcs, dce, height in CLEAR_RANGES_REL:
        for rr in range(top + drow, top + drow + height):
            for cc in range(left + dcs, left + dce + 1):
                if _is_protected(top, left, rr, cc):
                    continue
                safe_write(ws, rr, cc, "")

# ===== スロット数（テンプレ幅から右端まで） =====
def count_slots_in_template(ws) -> int:
    start = 2   # B列
    step  = 10  # 1スロット=10列
    maxcol = ws.max_column
    return max(1, min(30, (maxcol - start) // step + 1))

# ===== シートを1枚だけに整理し、月名でリネーム =====
def prepare_single_sheet(wb: openpyxl.Workbook, month:int) -> openpyxl.worksheet.worksheet.Worksheet:
    desired = month_sheet_name(month)
    if wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        for name in list(wb.sheetnames)[1:]:
            del wb[name]
        ws.title = desired
        return ws
    ws = wb.create_sheet(desired)
    return ws

# ===== クラスラベル =====
def write_class_label(ws, top:int, left:int, label:Optional[str]):
    safe_write(ws, top+1, left, label if label else "")

# ===== 出力（S/A/B 用：MAIN テンプレ） =====
def fill_sheet_main(ws_out, target_month:int, classes:dict, *, teacher_blank:bool=False):
    """
    classes = {'S':[Event...], 'A':[...], 'B':[...]}  ※Xは含めない
    各スロットでデータが無い段は、クラス欄ラベルも含めて空欄。

    追加仕様：
    - 同じスロット内で、いずれかのクラスの Event.special が True（「特」）の場合、
      そのスロットでは「特」のクラスだけを書き、他クラス（S/A/B）は完全に空欄のままにする。
    - その際、クラスラベルも「特」にする。
    """
    base_top = 6
    order = ["S","A","B"]
    idx = {k:i for i,k in enumerate(order)}
    total_slots = count_slots_in_template(ws_out)
    lens = {k: len(classes.get(k,[])) for k in order}

    for slot in range(total_slots):
        col_left = 2 + 10*slot

        # このスロットに載るイベントを一旦集める
        slot_events: Dict[str, Event] = {}
        has_special = False
        for k in order:
            if slot < lens[k]:
                ev = classes[k][slot]
                slot_events[k] = ev
                if getattr(ev, "special", False):
                    has_special = True

        # まずブロックをクリアし、クラスラベルを書く
        for k in order:
            top = base_top + ROW_STEP*idx[k]
            clear_one_block(ws_out, top, col_left)

            if slot >= lens[k]:
                # そもそも授業が無いスロット → ラベル空欄
                label = ""
            else:
                ev = slot_events[k]
                if has_special:
                    # 特スロット：特のクラスだけ「特」、その他はラベルも空欄
                    if getattr(ev, "special", False):
                        label = "特"
                    else:
                        label = ""
                else:
                    # 通常スロット：従来どおり S/A/B
                    label = k

            write_class_label(ws_out, top, col_left, label)

        # 書き込み用ヘルパ
        def _write_ev(top_row:int, ev:Event):
            safe_write(ws_out, top_row+4,  col_left, str(target_month) if target_month else "")
            safe_write(ws_out, top_row+5,  col_left, str(ev.day) if ev.day!="" else "")
            safe_write(ws_out, top_row+6,  col_left, ev.wday or "")
            safe_write(ws_out, top_row+8,  col_left, "" if teacher_blank else (ev.teacher or ""))

        # 実際のデータ書き込み
        for k in order:
            if slot >= lens[k]:
                continue
            ev = slot_events[k]
            top_row = base_top + ROW_STEP*idx[k]

            if has_special and not getattr(ev, "special", False):
                # 特がある縦列では、通常回は出力しない（空欄のまま）
                continue

            _write_ev(top_row, ev)

# ===== 出力（X 用：Xテンプレ。先頭段にX/特を書いていく） =====
def fill_sheet_x(ws_out, target_month:int, x_events:List[Event], *, teacher_blank:bool=False):
    base_top = 6
    total_slots = count_slots_in_template(ws_out)
    n = len(x_events)

    for slot in range(total_slots):
        col_left = 2 + 10*slot
        clear_one_block(ws_out, base_top, col_left)

        if slot < n:
            ev = x_events[slot]
            label = "特" if getattr(ev, "special", False) else "X"
        else:
            ev = None
            label = ""

        write_class_label(ws_out, base_top, col_left, label)

        if ev is not None:
            safe_write(ws_out, base_top+4,  col_left, str(target_month) if target_month else "")
            safe_write(ws_out, base_top+5,  col_left, str(ev.day) if ev.day!="" else "")
            safe_write(ws_out, base_top+6,  col_left, ev.wday or "")
            safe_write(ws_out, base_top+8,  col_left, "" if teacher_blank else (ev.teacher or ""))

# ===== 保存（PermissionError回避） =====
def _safe_save(wb, fname):
    p=Path(fname); stem,suf=p.stem,p.suffix or ".xlsx"
    for i in range(20):
        tgt=p if i==0 else p.with_name(f"{stem}({i}){suf}")
        try:
            wb.save(tgt); print(f"出力: {tgt.name}"); return
        except PermissionError:
            continue
    raise PermissionError(f"保存できません: {fname}")

# === 追加：科目表示名（小４〜６の数→算数） ===
def _display_subj_name(grade:str, subj:str) -> str:
    if subj == "数" and grade in ("4","5","6"):
        return "算数"
    return SUBJ_NAME[subj]

# ★★★ ファイル名に year, month を付ける ★★★
def save_main_book(wb, campus:str, grade:str, subj:str, year:int, month:int):
    subj_j = _display_subj_name(grade, subj)
    fname = f"{campus}{GRADE_LABEL[grade]}{subj_j}_{year}_{month:02d}.xlsx"
    _safe_save(wb, fname)

# ===== イベント収集 =====
def collect_events(sh, target_month:int)->List[Event]:
    legend = detect_teacher_legend(sh)
    room_row = detect_classroom_row(sh)
    C,AD = column_index_from_string("C"), column_index_from_string("AD")
    evs=[]; col_confirmed: Dict[int,str] = {}

    for r in range(2,55):
        day,w = parse_day_week_from_row(sh, r)
        for c in range(C,AD+1):
            txt=cell_text(sh.cell(r,c))
            if not txt: continue
            tok=parse_class_token(txt)
            if not tok: continue

            # 縦結合（2行固定）の下段救済：授業が見つかった行で日付/曜日が取れない場合のみ、直上1行も試す
            if day is None or w is None:
                day, w = parse_day_week_from_row(sh, r, allow_prev_row=True)

            g,k,subj,special=tok
            teacher = resolve_teacher(sh, legend, r, c, col_confirmed)
            time = find_timeband_above(sh, r, c)
            classroom = cell_text(sh.cell(room_row, c)) if room_row else ""
            evs.append(Event(target_month, day or "", w or "", time, classroom, g, k, subj, txt, teacher, special, r, c))

    evs.sort(key=lambda e: (int(e.day) if str(e.day).isdigit() else 99, e.r, e.c))
    return evs

# ===== 校舎単位の出力 =====
def export_one_campus(ws_schedule, campus:str, year:int, month:int):
    all_events = collect_events(ws_schedule, month)

    # (学年,科目,クラス) ごとに束ねる
    buckets=defaultdict(list)
    for e in all_events:
        buckets[(e.grade, e.subj, e.klass)].append(e)

    for grade in GRADES:
        for subj in SUBJS:
            s_list = buckets.get((grade, subj, "S"), [])
            a_list = buckets.get((grade, subj, "A"), [])
            b_list = buckets.get((grade, subj, "B"), [])
            x_list = buckets.get((grade, subj, "X"), [])

            # --- 本体: S/A/B を TEMPLATE_MAIN で出力 ---
            if any([s_list, a_list, b_list]):
                wb_main = openpyxl.load_workbook(TEMPLATE_MAIN)
                ws_out = prepare_single_sheet(wb_main, month)
                fill_sheet_main(ws_out, month, {"S":s_list,"A":a_list,"B":b_list}, teacher_blank=False)
                save_main_book(wb_main, campus, grade, subj, year, month)

            # --- X専用: TEMPLATE_X で出力 ---
            if x_list:
                wb_x = openpyxl.load_workbook(TEMPLATE_X)
                ws_x = prepare_single_sheet(wb_x, month)
                fill_sheet_x(ws_x, month, x_list, teacher_blank=False)
                subj_j = _display_subj_name(grade, subj)
                fname_x = f"{campus}{grade}X{subj_j}_{year}_{month:02d}.xlsx"
                _safe_save(wb_x, fname_x)

    # --- 補講（英語/数学）：TEMPLATE_MAIN を使用・先生欄空欄 ---
    for keyword, title in [("英語補講","英語補講"), ("数学補講","数学補講")]:
        hits=[e for e in all_events if keyword in e.text]
        if hits:
            wb = openpyxl.load_workbook(TEMPLATE_MAIN)
            ws = prepare_single_sheet(wb, month)
            fill_sheet_main(ws, month, {"S":hits,"A":[],"B":[]}, teacher_blank=True)
            fname_h = f"{campus}{title}_{year}_{month:02d}.xlsx"
            _safe_save(wb, fname_h)

# ===== シート選別 =====
def _norm(s:str)->str:
    return (s or "").replace(" ","").replace("\u3000","")

def choose_target_sheets(wb) -> List[Tuple[str, str, openpyxl.worksheet.worksheet.Worksheet]]:
    """
    戻り値: [(campus_label, sheetname, sheetobj), ...]
    campus_label は "本校" or "南教室"
    条件: シート名に「教務」を含み、かつ「本校」 or 「南/南教室」を含むもの
    """
    picks=[]
    for name in wb.sheetnames:
        nm=_norm(name)
        if "教務" not in nm: continue
        if "本校" in nm:
            picks.append(("本校", name, wb[name]))
        elif ("南教室" in nm) or (("南" in nm) and ("教室" in nm or "校" not in nm)):
            picks.append(("南教室", name, wb[name]))
    result=[]
    seen=set()
    for campus,name,sh in picks:
        if campus in seen: continue
        result.append((campus,name,sh)); seen.add(campus)
    return result

# ===== メイン =====
def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--prefer", type=str, help="優先月カンマ区切り（例: 11,10）")
    args=ap.parse_args()
    prefer=[int(x) for x in re.split(r"[,\s]+", args.prefer.strip()) if x] if args.prefer else None

    y,m,sch = pick_schedule_in_same_folder(prefer)
    print(f"[INFO] スケジュール: {sch.name} ({m}月)")
    wb_s = openpyxl.load_workbook(sch, data_only=True, keep_vba=True)

    targets = choose_target_sheets(wb_s)
    if not targets:
        print("対象シート（本校/南教室の教務部用）が見つかりません。")
        return

    for campus, sname, sh in targets:
        print(f"=== 校舎: {campus}（シート: {sname}） ===")
        export_one_campus(sh, campus, y, m)

if __name__ == "__main__":
    main()
