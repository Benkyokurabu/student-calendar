#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Zoom cloud recording URL importer for the lesson journal system.

Default mode is a dry run:
    python zoom_recording_urls.py --month 2026-08

Write matched recording URLs into OneDrive journal Excel files:
    python zoom_recording_urls.py --month 2026-08 --write

Required local secrets are read from .env or environment variables:
    ZOOM_ACCOUNT_ID
    ZOOM_CLIENT_ID
    ZOOM_CLIENT_SECRET
"""

from __future__ import annotations

import argparse
import base64
import calendar
import json
import os
import re
import shutil
import ssl
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
SYSTEM_DIR = SCRIPT_DIR.parent
MEETING_IDS_PATH = SCRIPT_DIR / "zoom_meeting_ids.json"
ENV_PATH = SCRIPT_DIR / ".env"
JST = timezone(timedelta(hours=9))

FIRST_BLOCK_COL = 2
BLOCK_WIDTH = 10
BASE_TOP_MAIN = {"S": 6, "A": 26, "B": 46}
BASE_TOP_X = 6

BACKUP_DIR_NAME = "_backup_zoom_recording_urls"


def load_dotenv(path: Path = ENV_PATH) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"{name} is not set. Create {ENV_PATH.name} from .env.example.")
    return value


def http_json(method: str, url: str, *, headers: Optional[dict] = None, data: Optional[bytes] = None) -> dict:
    req = urllib.request.Request(url, data=data, method=method, headers=headers or {})
    try:
        with urllib.request.urlopen(req, timeout=45, context=ssl.create_default_context()) as resp:
            body = resp.read().decode("utf-8")
            return json.loads(body) if body else {}
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Zoom API error {e.code}: {body}") from e


class ZoomClient:
    def __init__(self) -> None:
        load_dotenv()
        self.account_id = require_env("ZOOM_ACCOUNT_ID")
        self.client_id = require_env("ZOOM_CLIENT_ID")
        self.client_secret = require_env("ZOOM_CLIENT_SECRET")
        self._access_token: Optional[str] = None

    def access_token(self) -> str:
        if self._access_token:
            return self._access_token

        token = base64.b64encode(f"{self.client_id}:{self.client_secret}".encode("utf-8")).decode("ascii")
        qs = urllib.parse.urlencode({
            "grant_type": "account_credentials",
            "account_id": self.account_id,
        })
        resp = http_json(
            "POST",
            f"https://zoom.us/oauth/token?{qs}",
            headers={"Authorization": f"Basic {token}"},
        )
        access_token = str(resp.get("access_token", ""))
        if not access_token:
            raise RuntimeError("Zoom token response did not include access_token.")
        self._access_token = access_token
        return access_token

    def get_meeting_recordings(self, meeting_id: str) -> dict:
        clean_id = clean_meeting_id(meeting_id)
        url = f"https://api.zoom.us/v2/meetings/{urllib.parse.quote(clean_id)}/recordings"
        return http_json("GET", url, headers={"Authorization": f"Bearer {self.access_token()}"})

    def list_account_recordings(self, from_date: date, to_date: date) -> List[dict]:
        meetings: List[dict] = []
        next_page_token = ""
        while True:
            qs = {
                "page_size": "300",
                "from": from_date.isoformat(),
                "to": to_date.isoformat(),
            }
            if next_page_token:
                qs["next_page_token"] = next_page_token
            url = "https://api.zoom.us/v2/accounts/me/recordings?" + urllib.parse.urlencode(qs)
            payload = http_json("GET", url, headers={"Authorization": f"Bearer {self.access_token()}"})
            meetings.extend(payload.get("meetings") or [])
            next_page_token = str(payload.get("next_page_token") or "")
            if not next_page_token:
                return meetings

def normalize_digits(value: str) -> str:
    return value.translate(str.maketrans("０１２３４５６７８９：", "0123456789:"))


def clean_meeting_id(value: str) -> str:
    return re.sub(r"\D", "", str(value))

TIME_RE = re.compile(r"(\d{1,2}):(\d{2})")


def schedule_hour_to_24(hour: int) -> int:
    if 1 <= hour <= 11:
        return hour + 12
    return hour


def parse_lesson_window(ev: dict) -> Optional[Tuple[datetime, datetime]]:
    raw_date = str(ev.get("date", ""))
    raw_time = normalize_digits(str(ev.get("time", ""))).replace("~", "～")
    times = TIME_RE.findall(raw_time)
    if not raw_date or len(times) < 2:
        return None

    y, m, d = map(int, raw_date[:10].split("-"))
    sh, sm = map(int, times[0])
    eh, em = map(int, times[1])
    sh = schedule_hour_to_24(sh)
    eh = schedule_hour_to_24(eh)
    start = datetime.combine(date(y, m, d), time(sh, sm), JST)
    end = datetime.combine(date(y, m, d), time(eh, em), JST)
    if end <= start:
        end += timedelta(days=1)
    return start, end


def parse_zoom_time(value: str) -> Optional[datetime]:
    if not value:
        return None
    s = value.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(JST)


def recording_url(meeting: dict) -> str:
    files = meeting.get("recording_files") or []
    preferred = [
        "shared_screen_with_speaker_view",
        "shared_screen_with_gallery_view",
        "active_speaker",
        "gallery_view",
    ]
    for recording_type in preferred:
        for f in files:
            if str(f.get("status", "")).lower() != "completed":
                continue
            if f.get("recording_type") == recording_type and f.get("play_url"):
                return str(f["play_url"])
    for f in files:
        if str(f.get("status", "")).lower() == "completed" and f.get("play_url"):
            return str(f["play_url"])

    url = str(meeting.get("share_url") or "")
    passcode = str(meeting.get("recording_play_passcode") or "")
    if url and passcode and "pwd=" not in url:
        sep = "&" if "?" in url else "?"
        url = f"{url}{sep}pwd={urllib.parse.quote(passcode)}"
    return url


SUPPLEMENT_RE = re.compile(r"(英語|数学)\s*補講")


def supplement_title(ev: dict) -> Optional[str]:
    text = " ".join([
        str(ev.get("label") or ""),
        str(ev.get("displayTitle") or ""),
        str(ev.get("groupKey") or ""),
    ])
    match = SUPPLEMENT_RE.search(normalize_digits(text))
    if not match:
        return None
    return f"{match.group(1)}補講"


@dataclass
class RecordingCandidate:
    meeting_id: str
    start_time: datetime
    end_time: Optional[datetime]
    topic: str
    url: str
    raw: dict


def flatten_recordings(meeting_id: str, payload: dict) -> List[RecordingCandidate]:
    meetings = payload.get("meetings")
    if meetings is None:
        meetings = [payload]
    result: List[RecordingCandidate] = []
    for meeting in meetings:
        start = parse_zoom_time(str(meeting.get("start_time", "")))
        if start is None:
            continue
        duration = meeting.get("duration")
        end = start + timedelta(minutes=int(duration)) if isinstance(duration, int) else None
        url = recording_url(meeting)
        if not url:
            continue
        result.append(RecordingCandidate(
            meeting_id=meeting_id,
            start_time=start,
            end_time=end,
            topic=str(meeting.get("topic") or ""),
            url=url,
            raw=meeting,
        ))
    return result


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def load_schedule(month: str) -> List[dict]:
    path = SYSTEM_DIR / f"schedule_{month}.json"
    if not path.exists():
        path = SYSTEM_DIR / "schedule_latest.json"
    return load_json(path)


def determine_latest_schedule_month() -> str:
    events = load_json(SYSTEM_DIR / "schedule_latest.json")
    months = sorted({str(ev.get("date", ""))[:7] for ev in events if str(ev.get("date", ""))[:7]})
    if not months:
        return datetime.now(JST).strftime("%Y-%m")
    return months[-1]


def load_meeting_ids() -> Dict[str, Dict[str, str]]:
    return load_json(MEETING_IDS_PATH)


def meeting_id_for_event(ev: dict, meeting_ids: Dict[str, Dict[str, str]]) -> Optional[str]:
    campus = str(ev.get("campus", ""))
    room = str(ev.get("room", ""))
    return meeting_ids.get(campus, {}).get(room)


def relevant_events(events: Iterable[dict], month: str, meeting_ids: Dict[str, Dict[str, str]]) -> List[dict]:
    result = []
    for ev in events:
        if str(ev.get("date", ""))[:7] != month:
            continue
        if meeting_id_for_event(ev, meeting_ids):
            result.append(ev)
    return result


def month_date_range(month: str) -> Tuple[date, date]:
    year, month_num = map(int, month.split("-"))
    last_day = calendar.monthrange(year, month_num)[1]
    return date(year, month_num, 1), date(year, month_num, last_day)


def fetch_recordings_for_events(client: ZoomClient, events: List[dict], meeting_ids: Dict[str, Dict[str, str]], target_month: str) -> Dict[str, List[RecordingCandidate]]:
    unique_ids = sorted({clean_meeting_id(meeting_id_for_event(ev, meeting_ids) or "") for ev in events if meeting_id_for_event(ev, meeting_ids)})
    by_id: Dict[str, List[RecordingCandidate]] = {}
    for meeting_id in unique_ids:
        by_id[meeting_id] = []

    from_date, to_date = month_date_range(target_month)
    print(f"[zoom] account recordings {from_date.isoformat()} to {to_date.isoformat()}")
    try:
        meetings = client.list_account_recordings(from_date, to_date)
    except RuntimeError as e:
        raise RuntimeError(
            "Could not list account recordings. Check that the Server-to-Server OAuth app has "
            "Recording scopes such as View all user recordings. "
            f"Original error: {e}"
        ) from e

    for meeting in meetings:
        meeting_id = clean_meeting_id(str(meeting.get("id") or meeting.get("meeting_id") or ""))
        if meeting_id not in by_id:
            continue
        by_id[meeting_id].extend(flatten_recordings(meeting_id, {"meetings": [meeting]}))

    for meeting_id in unique_ids:
        by_id[meeting_id].sort(key=lambda r: r.start_time)
        print(f"[zoom] meeting {meeting_id}: {len(by_id[meeting_id])} recordings")
    return by_id


def match_recording(ev: dict, recordings: List[RecordingCandidate], tolerance_before: int, tolerance_after: int) -> Optional[RecordingCandidate]:
    window = parse_lesson_window(ev)
    if window is None:
        return None
    lesson_start, lesson_end = window
    search_start = lesson_start - timedelta(minutes=tolerance_before)
    search_end = lesson_start + timedelta(minutes=tolerance_after)
    candidates = [
        recording
        for recording in recordings
        if search_start <= recording.start_time <= search_end
        or (
            recording.end_time is not None
            and recording.start_time <= lesson_start
            and recording.end_time >= lesson_start
        )
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda r: abs((r.start_time - lesson_start).total_seconds()))
    return candidates[0]


def import_extract_helpers():
    sys.path.insert(0, str(SYSTEM_DIR))
    import extract_journal_to_json as ej
    return ej


def download_cloud_journal() -> Tuple[Path, Any]:
    sys.path.insert(0, str(SYSTEM_DIR))
    from download_journal_from_cloud import download_journal, upload_journal
    print("[cloud] downloading OneDrive journal files via rclone...")
    journal_dir = download_journal()
    return journal_dir, upload_journal


def get_default_journal_dir() -> Path:
    candidates = [
        Path.home() / "OneDrive" / "●勉強クラブ共有" / "09　授業日誌",
        Path(r"C:\Users\kudok\OneDrive\●勉強クラブ共有\09　授業日誌"),
    ]
    for c in candidates:
        if c.exists():
            return c
    raise FileNotFoundError("09　授業日誌 folder was not found.")


def merged_top_left_cell(ws, row: int, col: int):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col)
    return ws.cell(row, col)


def read_merged_text(ws, row: int, col: int) -> str:
    v = merged_top_left_cell(ws, row, col).value
    return "" if v is None else str(v).strip()


def write_merged_cell(ws, row: int, col: int, value: str) -> None:
    merged_top_left_cell(ws, row, col).value = value


def find_block_col_by_day(ws, top_row: int, day: int, want_special: Optional[bool]) -> Optional[int]:
    max_slots = max(1, min(40, (ws.max_column - FIRST_BLOCK_COL) // BLOCK_WIDTH + 1))
    matches: List[Tuple[int, bool]] = []
    for slot in range(max_slots):
        col = FIRST_BLOCK_COL + slot * BLOCK_WIDTH
        d = read_merged_text(ws, top_row + 5, col)
        if d and d == str(day):
            annual = ws.cell(row=2, column=col + 4).value
            is_special = isinstance(annual, str) and annual.strip() == "特"
            matches.append((col, is_special))
    if not matches:
        return None
    if want_special is not None:
        for col, is_special in matches:
            if is_special == want_special:
                return col
    return matches[0][0]


class WorkbookWriter:
    def __init__(self, journal_dir: Path, year: int, month: int) -> None:
        self.journal_dir = journal_dir
        self.year = year
        self.month = month
        self.ej = import_extract_helpers()
        self.cache: Dict[str, Tuple[Path, openpyxl.Workbook, Any]] = {}
        self.path_cache: Dict[str, Optional[Path]] = {}
        self.workbook_index: Optional[Dict[str, Path]] = None
        self.modified: set[str] = set()

    def workbook_filename(self, ev: dict) -> Optional[str]:
        supplement = supplement_title(ev)
        if supplement:
            campus = self.ej.CAMPUS_JP.get(str(ev.get("campus", "")))
            if not campus:
                return None
            return f"{campus}{supplement}_{self.year}.xlsx"
        return self.ej.workbook_filename_from_event(ev, self.year)


    def find_workbook_path(self, filename: str) -> Optional[Path]:
        if filename in self.path_cache:
            return self.path_cache[filename]
        direct = self.journal_dir / filename
        if direct.exists():
            self.path_cache[filename] = direct
            return direct
        if self.workbook_index is None:
            skip_dirs = getattr(self.ej, "SKIP_DIRS", set())
            self.workbook_index = {}
            for candidate in self.journal_dir.rglob("*.xlsx"):
                if not candidate.is_file():
                    continue
                try:
                    parts = set(candidate.relative_to(self.journal_dir).parts)
                except ValueError:
                    parts = set(candidate.parts)
                if skip_dirs & parts:
                    continue
                self.workbook_index.setdefault(candidate.name, candidate)
        path = self.workbook_index.get(filename)
        self.path_cache[filename] = path
        return path

    def get_top_row(self, klass: str) -> Optional[int]:
        if klass == "":
            return BASE_TOP_MAIN["S"]
        if klass == "X":
            return BASE_TOP_X
        return BASE_TOP_MAIN.get(klass)

    def get(self, ev: dict) -> Optional[Tuple[str, Path, openpyxl.Workbook, Any, int, int]]:
        filename = self.workbook_filename(ev)
        if not filename:
            return None
        if filename not in self.cache:
            path = self.find_workbook_path(filename)
            if path is None:
                return None
            wb = openpyxl.load_workbook(path, data_only=False, keep_vba=False)
            ws = self.ej.find_month_sheet(wb, self.year, self.month)
            if ws is None:
                wb.close()
                return None
            self.cache[filename] = (path, wb, ws)
        path, wb, ws = self.cache[filename]
        top_row = self.get_top_row(str(ev.get("class", "")))
        if top_row is None:
            return None
        day = int(str(ev.get("date", ""))[8:10])
        col = find_block_col_by_day(ws, top_row, day, ev.get("special") if "special" in ev else None)
        if col is None:
            return None
        return filename, path, wb, ws, top_row, col

    def existing_url(self, ev: dict) -> str:
        entry = self.get(ev)
        if entry is None:
            return ""
        _, _, _, ws, top_row, col = entry
        return read_merged_text(ws, top_row + 11, col + 2)

    def write_url(self, ev: dict, url: str) -> bool:
        entry = self.get(ev)
        if entry is None:
            return False
        filename, _, _, ws, top_row, col = entry
        write_merged_cell(ws, top_row + 11, col + 2, url)
        self.modified.add(filename)
        return True

    def save_all(self, dry_run: bool) -> None:
        if dry_run:
            return
        backup_root = self.journal_dir / BACKUP_DIR_NAME / datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_root.mkdir(parents=True, exist_ok=True)
        for filename in sorted(self.modified):
            path, wb, _ = self.cache[filename]
            shutil.copy2(path, backup_root / path.name)
            fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
            os.close(fd)
            tmp = Path(tmp_name)
            try:
                wb.save(tmp)
                shutil.copy2(tmp, path)
                print(f"[write] {filename}")
            finally:
                if tmp.exists():
                    tmp.unlink()

    def close_all(self) -> None:
        for _, wb, _ in self.cache.values():
            try:
                wb.close()
            except Exception:
                pass


def event_label(ev: dict) -> str:
    return " ".join([
        str(ev.get("date", "")),
        str(ev.get("time", "")),
        str(ev.get("campus", "")),
        f"room{ev.get('room', '')}",
        str(ev.get("label") or ev.get("groupKey") or ""),
    ]).strip()


def main() -> int:
    ap = argparse.ArgumentParser(description="Fetch Zoom cloud recording URLs and write them into lesson journal Excel files.")
    ap.add_argument("--month", help="Target month, e.g. 2026-08. Defaults to the latest month in schedule_latest.json.")
    ap.add_argument("--write", action="store_true", help="Actually write URLs into OneDrive journal Excel files.")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing recording URLs.")
    ap.add_argument("--check-existing", action="store_true", help="In dry-run mode, open journal Excel files and skip rows that already have URLs.")
    ap.add_argument("--journal-dir", help="Override journal folder path.")
    ap.add_argument("--cloud", action="store_true", help="Download journal files from OneDrive cloud with rclone and upload changes after writing.")
    ap.add_argument("--graph", action="store_true", help="Write recording URLs directly to Excel Online cells with Microsoft Graph.")
    ap.add_argument("--tolerance-before", type=int, default=30, help="Minutes before lesson start to accept a recording.")
    ap.add_argument("--tolerance-after", type=int, default=30, help="Minutes after lesson start to accept a recording.")
    args = ap.parse_args()
    if args.cloud and args.journal_dir:
        raise RuntimeError("--cloud and --journal-dir cannot be used together.")
    if args.graph and (args.cloud or args.journal_dir):
        raise RuntimeError("--graph cannot be used with --cloud or --journal-dir.")

    target_month = args.month or determine_latest_schedule_month()
    year, month = map(int, target_month.split("-"))
    meeting_ids = load_meeting_ids()
    events = relevant_events(load_schedule(target_month), target_month, meeting_ids)
    print(f"[schedule] {target_month}: {len(events)} events have mapped Zoom meeting IDs")

    client = ZoomClient()
    recordings_by_id = fetch_recordings_for_events(client, events, meeting_ids, target_month)

    writer = None
    upload_journal = None
    if args.write or args.check_existing:
        if args.graph:
            from graph_excel_writer import GraphWorkbookWriter
            helper = WorkbookWriter(Path("."), year, month)
            writer = GraphWorkbookWriter(target_month, helper.workbook_filename)
        elif args.cloud:
            journal_dir, upload_journal = download_cloud_journal()
            writer = WorkbookWriter(journal_dir, year, month)
        else:
            journal_dir = Path(args.journal_dir) if args.journal_dir else get_default_journal_dir()
            writer = WorkbookWriter(journal_dir, year, month)

    matched = 0
    skipped_existing = 0
    missing = 0
    write_failed = 0
    try:
        for ev in events:
            meeting_id = meeting_id_for_event(ev, meeting_ids)
            rec = match_recording(ev, recordings_by_id.get(clean_meeting_id(meeting_id or ""), []), args.tolerance_before, args.tolerance_after)
            if rec is None:
                missing += 1
                continue
            if writer is not None:
                existing = writer.existing_url(ev)
                if existing and not args.overwrite:
                    skipped_existing += 1
                    continue
            matched += 1
            print(f"[match] {event_label(ev)} -> {rec.start_time.strftime('%Y-%m-%d %H:%M')} {rec.url}")
            if args.write:
                if not writer.write_url(ev, rec.url):
                    write_failed += 1
                    print(f"[WARN] Excel target not found: {event_label(ev)}")
        if writer is not None:
            writer.save_all(dry_run=not args.write)
            if args.write and upload_journal is not None and writer.modified:
                print("[cloud] uploading changed journal files via rclone...")
                upload_journal(journal_dir)
    finally:
        if writer is not None:
            writer.close_all()

    mode = "WRITE" if args.write else "DRY-RUN"
    print(f"[done] mode={mode} matched={matched} skipped_existing={skipped_existing} missing={missing} write_failed={write_failed}")
    if not args.write:
        print("[next] Add --write to update Excel files after reviewing matches.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except RuntimeError as e:
        print(f"[ERROR] {e}")
        raise SystemExit(1)


















