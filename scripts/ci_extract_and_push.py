#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GitHub Actions (CI) 用パイプライン。
rclone でダウンロード済みのファイルを使って JSON を生成する。
git commit/push はワークフロー側で行うため、ここでは行わない。
"""

import calendar
import json
import subprocess
import sys
import os
from datetime import date
from pathlib import Path
import shutil


def get_months_to_process(month_arg=None, *, script_dir=None):
    if month_arg:
        return [month_arg]
    today = date.today()
    last_day = calendar.monthrange(today.year, today.month)[1]
    current_month = f"{today.year}-{today.month:02d}"

    months = []

    if today.day > last_day - 5:
        if today.month == 1:
            prev_month = f"{today.year - 1}-12"
        else:
            prev_month = f"{today.year}-{today.month - 1:02d}"
        months.append(prev_month)

    months.append(current_month)

    # 来月のスケジュールExcelが存在すれば来月も処理対象に含める
    if today.month == 12:
        next_year, next_month = today.year + 1, 1
    else:
        next_year, next_month = today.year, today.month + 1
    next_month_str = f"{next_year}-{next_month:02d}"
    if next_month_str not in months and script_dir:
        import re as _re
        # CI環境: _backup/スケジュール表/ にスケジュールExcelがある
        backup_dir = script_dir / "_cloud_journal" / "_backup" / "スケジュール表"
        if backup_dir.exists():
            pattern = _re.compile(rf"^{next_year}年0?{next_month}月スケジュール\.xlsm$")
            if any(pattern.match(f.name) for f in backup_dir.iterdir() if f.is_file()):
                months.append(next_month_str)
        # フォールバック: script_dir直下も確認
        if next_month_str not in months:
            pattern = _re.compile(rf"^{next_year}年0?{next_month}月スケジュール\.xlsm$")
            if any(pattern.match(f.name) for f in script_dir.iterdir() if f.is_file()):
                months.append(next_month_str)

    return months


SKIP_DIRS = {"_backup", "退避", "__pycache__"}


def find_workbook_in_journal(journal_dir: Path, filename: str):
    p = journal_dir / filename
    if p.exists():
        return p
    for candidate in journal_dir.rglob(filename):
        if candidate.is_file():
            if SKIP_DIRS & {part for part in candidate.relative_to(journal_dir).parts}:
                continue
            return candidate
    return None


def create_month_sheets(script_dir: Path, months: list, journal_dir: Path):
    """スケジュールExcelを読み、日誌ファイルに新しい月シートを追加する"""
    old_cwd = os.getcwd()
    try:
        sys.path.insert(0, str(script_dir))
        os.chdir(str(script_dir))

        from export_by_grade_subject import (
            GRADES, SUBJS, GRADE_LABEL, _display_subj_name,
            pick_schedule_in_same_folder, choose_target_sheets, collect_events,
            create_month_sheet, set_header_cells,
            fill_sheet_main, fill_sheet_x, save_year_workbook,
            open_or_create_year_workbook, ensure_hidden_template_sheet,
            TEMPLATE_MAIN, TEMPLATE_X,
        )
        from collections import defaultdict
        import openpyxl

        created_count = 0

        for month_str in months:
            year, month = int(month_str.split("-")[0]), int(month_str.split("-")[1])

            try:
                y, m, sch = pick_schedule_in_same_folder([month])
            except FileNotFoundError:
                print(f"  [SKIP] {month}月のスケジュールが見つかりません")
                continue

            if m != month:
                print(f"  [SKIP] {month}月のスケジュールが見つかりません（{m}月のみ）")
                continue

            wb_s = openpyxl.load_workbook(sch, data_only=True, keep_vba=True)
            targets = choose_target_sheets(wb_s)
            if not targets:
                print(f"  [SKIP] 教務部用シートが見つかりません")
                continue

            for campus, sname, sh in targets:
                all_events = collect_events(sh, month, campus, year, month)
                buckets = defaultdict(list)
                for e in all_events:
                    if ("英語補講" in e.text) or ("数学補講" in e.text):
                        continue
                    buckets[(e.grade, e.subj, e.klass)].append(e)

                for grade in GRADES:
                    for subj in SUBJS:
                        s_list = buckets.get((grade, subj, "S"), [])
                        a_list = buckets.get((grade, subj, "A"), [])
                        b_list = buckets.get((grade, subj, "B"), [])
                        x_list = buckets.get((grade, subj, "X"), [])

                        subj_j = _display_subj_name(grade, subj)
                        grade_j = GRADE_LABEL[grade]

                        if any([s_list, a_list, b_list]):
                            fname = f"{campus}{grade_j}{subj_j}_{year}.xlsx"
                            wb_path = find_workbook_in_journal(journal_dir, fname)
                            if wb_path is None:
                                continue
                            wb = open_or_create_year_workbook(wb_path, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")
                            tmpl = ensure_hidden_template_sheet(wb, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")
                            ws_out = create_month_sheet(wb, tmpl, year, month)
                            if ws_out is None:
                                continue
                            set_header_cells(ws_out, campus, grade_j, subj_j, month, wb=wb, year=year)
                            fill_sheet_main(ws_out, month, {"S": s_list, "A": a_list, "B": b_list}, teacher_blank=False)
                            save_year_workbook(wb, wb_path)
                            print(f"  [NEW] {fname} + {ws_out.title}")
                            created_count += 1

                        if x_list:
                            fname = f"{campus}{grade_j}X{subj_j}_{year}.xlsx"
                            wb_path = find_workbook_in_journal(journal_dir, fname)
                            if wb_path is None:
                                continue
                            wb = open_or_create_year_workbook(wb_path, TEMPLATE_X, "__TEMPLATE_X__")
                            tmpl = ensure_hidden_template_sheet(wb, TEMPLATE_X, "__TEMPLATE_X__")
                            ws_out = create_month_sheet(wb, tmpl, year, month)
                            if ws_out is None:
                                continue
                            set_header_cells(ws_out, campus, grade_j, subj_j, month, wb=wb, year=year)
                            fill_sheet_x(ws_out, month, x_list, teacher_blank=False)
                            save_year_workbook(wb, wb_path)
                            print(f"  [NEW] {fname} + {ws_out.title}")
                            created_count += 1

                for keyword, title in [("英語補講", "英語補講"), ("数学補講", "数学補講")]:
                    hits = [e for e in all_events if keyword in e.text]
                    if not hits:
                        continue
                    fname = f"{campus}{title}_{year}.xlsx"
                    wb_path = find_workbook_in_journal(journal_dir, fname)
                    if wb_path is None:
                        continue
                    wb = open_or_create_year_workbook(wb_path, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")
                    tmpl = ensure_hidden_template_sheet(wb, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")
                    ws_out = create_month_sheet(wb, tmpl, year, month)
                    if ws_out is None:
                        continue
                    set_header_cells(ws_out, campus, "", title, month, wb=wb, year=year)
                    fill_sheet_main(ws_out, month, {"S": hits, "A": [], "B": []}, teacher_blank=True)
                    save_year_workbook(wb, wb_path)
                    print(f"  [NEW] {fname} + {ws_out.title}")
                    created_count += 1

            wb_s.close()

        os.chdir(old_cwd)
        return created_count

    except Exception as e:
        os.chdir(old_cwd)
        print(f"  [ERROR] 月シート作成中にエラー: {e}")
        import traceback
        traceback.print_exc()
        return 0


def run(args, **kwargs):
    print(f"  > {' '.join(str(a) for a in args)}")
    result = subprocess.run(args, **kwargs)
    if result.returncode != 0:
        print(f"[ERROR] コマンド失敗 (exit {result.returncode})")
        sys.exit(1)
    return result


def main():
    script_dir = Path(__file__).parent.resolve()
    repo_dir = script_dir.parent  # リポのルート
    month_arg = sys.argv[1] if len(sys.argv) > 1 else None
    months = get_months_to_process(month_arg, script_dir=script_dir)

    journal_dir = script_dir / "_cloud_journal"
    if not journal_dir.exists():
        print("[ERROR] _cloud_journal フォルダがありません。ワークフローでダウンロードしてください。")
        sys.exit(1)

    print("=" * 48)
    print("  CI: スケジュール＋授業日誌 JSON 生成")
    print(f"  対象月: {', '.join(months)}")
    print(f"  日誌フォルダ: {journal_dir}")
    print("=" * 48)
    print()

    # --- 1. スケジュールJSON再生成 ---
    print("[1/3] スケジュールExcelから JSON を生成中...")
    export_script = script_dir / "export_schedule_json.py"
    if export_script.exists():
        # months に含まれる月ごとに該当スケジュールファイルを明示指定して生成する
        # （引数なし実行だと今月優先で1ファイルしか処理されず、来月分が生成されないため）
        import re as _re_exp
        _RE_SCH_EXP = _re_exp.compile(r"^(\d{4})年0?(\d{1,2})月スケジュール\.xlsm$")
        for m in months:
            month_num = int(m.split("-")[1])
            sch_path = None
            for p in script_dir.iterdir():
                mm = _RE_SCH_EXP.match(p.name)
                if mm and int(mm.group(2)) == month_num:
                    sch_path = p
                    break
            if sch_path is None:
                print(f"  [SKIP] {m}のスケジュールが見つかりません")
                continue
            run([sys.executable, str(export_script), "--schedule", str(sch_path.resolve())],
                cwd=str(script_dir))
        for f in script_dir.glob("schedule_*.json"):
            dst = repo_dir / f.name
            shutil.copy2(f, dst)
            print(f"  → {dst.name}")
    else:
        print("[SKIP] export_schedule_json.py が見つかりません")
    print()

    # --- 1.1. journal_map 再生成 ---
    print("[1.1] journal_map を生成中...")
    build_map_script = script_dir / "build_journal_map.py"
    if build_map_script.exists():
        for m in months:
            month_num = m.split("-")[1]
            result = subprocess.run(
                [sys.executable, str(build_map_script), "--prefer", month_num],
                cwd=str(script_dir),
                capture_output=True, text=True, encoding="utf-8", errors="replace"
            )
            if result.stdout:
                for line in result.stdout.strip().split("\n"):
                    print(f"  {line}")
            if result.returncode != 0:
                print(f"  [WARN] journal_map 生成失敗 ({m}): {result.stderr.strip()}")
        # 生成されたjournal_mapをリポにコピー
        for f in script_dir.glob("journal_map_*.json"):
            dst = repo_dir / f.name
            shutil.copy2(f, dst)
            print(f"  → {dst.name}")
    else:
        print("[SKIP] build_journal_map.py が見つかりません。")
    print()

    # --- 1.5. 月シート作成（必要な場合） ---
    print("[1.5] 授業日誌Excelに新しい月シートを追加中...")
    created = create_month_sheets(script_dir, months, journal_dir)
    if created > 0:
        print(f"  → {created} 個のシートを作成しました")
    else:
        print("  → 新しいシートはありません（既に作成済み）")
    print()

    # --- 1.55. 既存シートの連鎖数式を修復 ---
    print("[1.55] 授業回数の連鎖数式を検証・修復中...")
    try:
        sys.path.insert(0, str(script_dir))
        from export_by_grade_subject import (count_slots_in_template, patch_counter_formulas,
            mark_special_counters, gray_out_block, clear_gray_block, _slot_has_day)
        import openpyxl as _openpyxl
        import re as _re_sheet
        _RE_MONTH_SHEET = _re_sheet.compile(r"^(\d{4})-(\d{2})$")

        formula_repair_count = 0
        special_restore_count = 0
        for xlsx_path in sorted(journal_dir.rglob("*.xlsx")):
            if SKIP_DIRS & {p for p in xlsx_path.relative_to(journal_dir).parts}:
                continue
            if xlsx_path.name.startswith("~"):
                continue
            try:
                wb = _openpyxl.load_workbook(xlsx_path)
            except Exception:
                continue

            month_sheets = []
            for sn in wb.sheetnames:
                _m = _RE_MONTH_SHEET.match(sn)
                if _m:
                    month_sheets.append((int(_m.group(1)), int(_m.group(2)), sn))
            month_sheets.sort()

            changed = False
            for _y, _mo, sn in month_sheets:
                ws = wb[sn]

                _total = count_slots_in_template(ws)
                for _si in range(_total):
                    _cl = 2 + 10 * _si
                    _labels = [ws.cell(row=r, column=_cl).value for r in (7, 27, 47)]
                    if any(isinstance(lb, str) and lb.strip() == "特" for lb in _labels):
                        _f2_val = ws.cell(row=2, column=_cl + 4).value
                        if not (isinstance(_f2_val, str) and _f2_val.strip() == "特"):
                            mark_special_counters(ws, _cl)
                            changed = True
                            special_restore_count += 1

                patch_counter_formulas(ws)
                changed = True
                formula_repair_count += 1

                is_x = any(
                    isinstance(ws.cell(row=7, column=2 + 10 * s).value, str)
                    and ws.cell(row=7, column=2 + 10 * s).value.strip() == "X"
                    for s in range(_total) if _slot_has_day(ws, 2 + 10 * s)
                )
                last_used = -1
                for s in range(_total):
                    if _slot_has_day(ws, 2 + 10 * s):
                        last_used = s
                if is_x:
                    base_top = 6
                    for s in range(_total):
                        cl = 2 + 10 * s
                        if s <= last_used:
                            clear_gray_block(ws, base_top, cl)
                        else:
                            gray_out_block(ws, base_top, cl)
                else:
                    order_tops = [6, 26, 46]
                    for s in range(_total):
                        cl = 2 + 10 * s
                        for top in order_tops:
                            if s <= last_used:
                                clear_gray_block(ws, top, cl)
                            else:
                                gray_out_block(ws, top, cl)

            if changed:
                wb.save(xlsx_path)
            wb.close()

        if special_restore_count > 0:
            print(f"  → 特マーク復元: {special_restore_count} 個")
        print(f"  → 連鎖数式再適用: {formula_repair_count} 個のシート")
        if special_restore_count == 0 and formula_repair_count == 0:
            print("  → すべて正常です")
    except Exception as e:
        print(f"  [ERROR] 回数修復中にエラー: {e}")
        import traceback
        traceback.print_exc()
    print()

    # --- 1.6. 既存シートの日付・講師を最新スケジュールに同期 ---
    print("[1.6] 既存シートの日付・講師をスケジュールに同期中...")
    import re as _re_sch
    _RE_SCH = _re_sch.compile(r"^(\d{4})年0?(\d{1,2})月スケジュール\.xlsm$")
    update_journal_py = repo_dir / "update_journal.py"
    if update_journal_py.exists():
        for m in months:
            month_num = int(m.split("-")[1])
            sch_path = None
            for p in script_dir.iterdir():
                _mm = _RE_SCH.match(p.name)
                if _mm and int(_mm.group(2)) == month_num:
                    sch_path = p
                    break
            if sch_path is None:
                print(f"  [SKIP] {month_num}月のスケジュールが見つかりません")
                continue
            print(f"  → {m} ({sch_path.name})")
            _result = subprocess.run(
                [sys.executable, str(update_journal_py),
                 "--schedule", str(sch_path.resolve()),
                 "--journal-dir", str(journal_dir)],
                cwd=str(repo_dir),
                capture_output=True, text=True, encoding="utf-8", errors="replace"
            )
            if _result.stdout:
                for line in _result.stdout.strip().split("\n"):
                    print(f"    {line}")
            if _result.returncode != 0:
                print(f"  [WARN] update_journal 失敗 (exit {_result.returncode})")
                if _result.stderr:
                    print(f"    {_result.stderr.strip()}")
    else:
        print("  [SKIP] update_journal.py が見つかりません")
    print()

    # --- 2. 日誌キャンパス間コピー + JSON抽出 ---
    # 日誌JSON退避（マージ用）
    old_jsons = {}
    for m in months:
        repo_journal = repo_dir / f"journal_{m}.json"
        if repo_journal.exists():
            old_jsons[m] = json.loads(repo_journal.read_text(encoding="utf-8"))

    for m in months:
        print(f"--- {m} の処理 ---")

        # 日誌キャンパス間コピー
        print(f"[2/3] オンラインペア授業の日誌をキャンパス間でコピー中... ({m})")
        sync_script = script_dir / "sync_journal_across_campus.py"
        if sync_script.exists():
            run([sys.executable, str(sync_script), "--month", m,
                 "--journal-dir", str(journal_dir),
                 "--repo-dir", str(repo_dir)])
        print()

        # 授業日誌JSON抽出
        print(f"[3/3] 授業日誌Excelから JSON を抽出中... ({m})")
        run([sys.executable, str(script_dir / "extract_journal_to_json.py"),
             "--month", m, "--journal-dir", str(journal_dir),
             "--repo-dir", str(repo_dir)])
        print()

    # --- マージ ---
    FIELDS = ["content", "page", "report", "recordingUrl",
              "teacher", "absence", "curriculumProgress", "note",
              "sessionNumber", "monthNum", "weekNum"]

    def entry_has_data(entry):
        if not isinstance(entry, dict):
            return False
        return any(entry.get(f) for f in FIELDS) or bool(entry.get("homework"))

    def merge_entry(old_entry, new_entry):
        merged = {}
        for f in FIELDS:
            new_val = new_entry.get(f, "")
            old_val = old_entry.get(f, "")
            merged[f] = new_val if new_val else old_val
        new_hw = new_entry.get("homework", [])
        old_hw = old_entry.get("homework", [])
        merged["homework"] = new_hw if new_hw else old_hw
        # prevEntry はマージ対象外（常に新しい方を採用）
        if "prevEntry" in new_entry:
            merged["prevEntry"] = new_entry["prevEntry"]
        return merged

    print("[MERGE] 日誌データのマージ中...")
    for m in months:
        if m not in old_jsons:
            print(f"  {m}: 前回データなし（初回生成）— そのまま採用")
            continue
        repo_journal = repo_dir / f"journal_{m}.json"
        if not repo_journal.exists():
            continue
        try:
            new_data = json.loads(repo_journal.read_text(encoding="utf-8"))
            old_data = old_jsons[m]
            new_entries = new_data.get("entries", {})
            old_entries = old_data.get("entries", {})

            merged_entries = {}
            all_keys = set(list(old_entries.keys()) + list(new_entries.keys()))
            kept_count = 0
            updated_count = 0

            for key in all_keys:
                old_val = old_entries.get(key, {})
                new_val = new_entries.get(key, {})
                if not isinstance(old_val, dict):
                    old_val = {}
                if not isinstance(new_val, dict):
                    new_val = {}

                if entry_has_data(new_val) and entry_has_data(old_val):
                    merged_entries[key] = merge_entry(old_val, new_val)
                    updated_count += 1
                elif entry_has_data(new_val):
                    merged_entries[key] = new_val
                    updated_count += 1
                elif entry_has_data(old_val):
                    merged_entries[key] = old_val
                    kept_count += 1
                else:
                    merged_entries[key] = new_val if key in new_entries else old_val

            if kept_count > 0:
                print(f"  {m}: {updated_count}件更新, {kept_count}件前回データ維持")
            else:
                print(f"  {m}: OK（{updated_count}件更新）")

            new_data["entries"] = merged_entries
            _text = json.dumps(new_data, ensure_ascii=False, indent=2)
            _tmp = repo_journal.with_suffix(".json.tmp")
            _tmp.write_text(_text, encoding="utf-8")
            _tmp.replace(repo_journal)
            latest = repo_dir / "journal_latest.json"
            _tmp2 = latest.with_suffix(".json.tmp")
            _tmp2.write_text(_text, encoding="utf-8")
            _tmp2.replace(latest)
        except Exception as e:
            print(f"[WARNING] マージ中にエラー: {e}")
    print()

    # --- 18か月より古い journal/schedule JSON を削除 ---
    import re as _re2
    _today = date.today()
    _cutoff_y, _cutoff_m = _today.year, _today.month - 18
    while _cutoff_m <= 0:
        _cutoff_y -= 1
        _cutoff_m += 12
    cutoff = f"{_cutoff_y:04d}-{_cutoff_m:02d}"
    removed = []
    for pat in ["journal_2*-*.json", "schedule_2*-*.json", "journal_map_2*-*.json"]:
        for f in repo_dir.glob(pat):
            m = _re2.search(r"(\d{4}-\d{2})", f.name)
            if m and m.group(1) < cutoff:
                f.unlink()
                removed.append(f.name)
    if removed:
        print(f"[CLEAN] 18か月超の古いJSONを削除: {', '.join(removed)}")
    print()

    print("=" * 48)
    print("  CI パイプライン完了（JSON生成済み）")
    print("=" * 48)


if __name__ == "__main__":
    main()
