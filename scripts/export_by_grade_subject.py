# -*- coding: utf-8 -*-
"""
授業日誌自動生成（統合版／年ファイルに月タブ追記）— 講師推測なし＆破損xlsx厳格対応

【今回の重要変更（あなたの優先事項）】
- 講師判定は「色が凡例と一致したときだけ」確定します。
  一致しない場合は空欄にします（= 間違った講師名を入れない）。
- 一致しないセルは txt に一覧出力します（未判定講師_校舎_YYYY-MM.txt）。

【破損xlsxの扱い】
- 既存の出力xlsxが壊れていた場合は __broken へ退避（リネーム）します。
- 退避できない（=ファイルロック）場合は、そこで停止して手動対応を促します。
  ※中途半端に別名出力して運用が分岐するのを防ぐため。

【確定仕様（このチャットで決定）】
- 先生凡例（色→先生名）：old方式（H56:AF56 → 54-58行(C〜AZ) → 40-120行(C〜AZ) 自動検出）
- 日付/曜日：new方式（結合セル左上＋縦結合下段救済）
- 出力：1年＝1ファイル、月ごとにタブ追加（YYYY-MM）
- 既に同じ月タブがある場合は、(再1) を作らず「既にあります」と表示してスキップ
- 新しい月タブは右端に追加
- ファイル名：年のみ付与（_YYYY）
- 「特」：授業名に「特」が含まれる回
  - 同一スロットに特がある場合：特のみ表示、他クラスはラベル含め完全空欄
  - 特ラベルは「特」
  - 特スロットは回数セル（2行目：年回数、3行目：月内回数）に「特」を表示
  - 月セル（$E$3参照）は触らない
- X：通常「X」、X特は「特」
- 補講：英語補講／数学補講ファイルも出力（年ファイル＋月タブ）
- 重複授業は除去
- 重複割当チェック（同日同時間同講師が複数）を txt に出す

【必要ファイル（同フォルダ）】
- YYYY年M月スケジュール.xlsm
- TEMPLATE_MAIN.xlsx（S/A/B用）
- TEMPLATE_X.xlsx（X用）

【実行】
python export_by_grade_subject.py
"""

import sys
import re
import argparse
import math
import zipfile
from pathlib import Path
from collections import defaultdict, namedtuple
from typing import Optional, Dict, Tuple, Set, List, Any
from copy import copy

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import MergedCell

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ===== 同フォルダのテンプレ =====
TEMPLATE_MAIN = Path("TEMPLATE_MAIN.xlsx")
TEMPLATE_X    = Path("TEMPLATE_X.xlsx")

# ===== 定数 =====
ROW_STEP = 20
GRADES = ["1", "2", "3", "4", "5", "6"]
SUBJS  = ["英", "数", "国", "理", "社"]

GRADE_LABEL = {"1": "中１", "2": "中２", "3": "中３", "4": "小４", "5": "小５", "6": "小６"}
SUBJ_NAME   = {"英": "英語", "数": "数学", "国": "国語", "理": "理科", "社": "社会"}

# 講師推測はしない（安全第一）
STRICT_TEACHER = True
ALLOW_NEAR_RGB = False  # 近傍RGBも使わない（誤判定防止）

RGB_NEAR_THRESH = 24

Event = namedtuple("Event", "month day wday time classroom grade klass subj text teacher special r c")

# ===== スケジュール自動選択 =====
RE_SCHEDULE = re.compile(r"^(\d{4})年0?(\d{1,2})月スケジュール\.xlsm$", re.ASCII)

def pick_schedule_in_same_folder(prefer=None):
    here = Path(".")
    cands = []
    for p in here.iterdir():
        if not p.is_file():
            continue
        m = RE_SCHEDULE.match(p.name)
        if m:
            cands.append((int(m.group(1)), int(m.group(2)), p))

    if not cands:
        raise FileNotFoundError("同フォルダに『YYYY年M月スケジュール.xlsm』が見つかりません。")

    if prefer:
        for pm in prefer:
            for (y, mo, pp) in sorted(cands, key=lambda x: (x[0], x[1]), reverse=True):
                if mo == pm:
                    return y, mo, pp

    y, mo, pp = sorted(cands, key=lambda x: (x[0], x[1]), reverse=True)[0]
    return y, mo, pp

# ===== 月タブ名 =====
def month_sheet_base_name(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"

# ===== ユーティリティ =====
def cell_text(c):
    v = c.value
    return "" if v is None else str(v).strip()

def normalize_digits(s):
    return s.translate(str.maketrans({chr(ord('０') + i): str(i) for i in range(10)}))

def looks_like_name(s: str) -> bool:
    return bool(re.fullmatch(r"[ぁ-んァ-ン一-龥々〆ヵヶA-Za-z]{1,6}", s or ""))

def parse_class_token(txt: str):
    s = normalize_digits(txt).replace(" ", "").replace("\u3000", "")
    if any(x in s for x in ("休講", "休み", "休校", "休", "テスト", "模試")):
        return None
    m = re.search(r"([1-6])\s*([SABXＳＡＢＸ])?(特)?\s*([数算国英理社])", s)
    if not m:
        return None
    grade = m.group(1)
    klass = (m.group(2) or "").upper().replace("Ｓ", "S").replace("Ａ", "A").replace("Ｂ", "B").replace("Ｘ", "X")
    special = bool(m.group(3))
    subj = m.group(4)
    subj_norm = "数" if subj in ("数", "算") else subj
    return grade, klass, subj_norm, special

def safe_write(ws, row, col, val):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            ws.cell(rng.min_row, rng.min_col).value = val
            return
    ws.cell(row=row, column=col).value = val

def merged_top_left_cell(sh, row: int, col: int):
    for mr in sh.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return sh.cell(mr.min_row, mr.min_col)
    return sh.cell(row, col)

def merged_value_text(sh, row: int, col: int) -> str:
    c = merged_top_left_cell(sh, row, col)
    return cell_text(c)

# ===== 色キー（塗りのみ：rgb / indexed / theme(+tint)） =====
def _rgb_from_argb(a: Optional[str]) -> Optional[str]:
    if not a:
        return None
    h = a.upper()
    if len(h) == 8:
        h = h[2:]
    return h if len(h) == 6 else None

def _fill_color_keys(fill) -> Set[Tuple[str, Any]]:
    keys = set()
    for col in (getattr(fill, "start_color", None), getattr(fill, "fgColor", None)):
        if not col:
            continue
        t = getattr(col, "type", None)
        if t == "rgb" and getattr(col, "rgb", None):
            rgb = _rgb_from_argb(col.rgb)
            if rgb:
                keys.add(("rgb", rgb))
        elif t == "indexed" and getattr(col, "indexed", None) is not None:
            keys.add(("indexed", col.indexed))
        elif t == "theme" and getattr(col, "theme", None) is not None:
            tint = getattr(col, "tint", 0.0)
            try:
                tint_val = float(tint)
            except Exception:
                tint_val = 0.0
            keys.add(("theme", col.theme, round(tint_val, 3)))
    return keys

def _hex_to_rgb_tuple(h: str) -> Tuple[int, int, int]:
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

def _rgb_dist(h1: str, h2: str) -> float:
    r1, g1, b1 = _hex_to_rgb_tuple(h1)
    r2, g2, b2 = _hex_to_rgb_tuple(h2)
    return math.sqrt((r1 - r2) ** 2 + (g1 - g2) ** 2 + (b1 - b2) ** 2)

# ===== 先生凡例（old方式：固定→補完→自動検出） =====
def detect_teacher_legend(ws) -> Dict[Tuple[str, Any], str]:
    mapping: Dict[Tuple[str, Any], str] = {}

    H, AF = column_index_from_string("H"), column_index_from_string("AF")
    C, AZ = column_index_from_string("C"), column_index_from_string("AZ")

    def _add_row(rr: int, c1: int, c2: int):
        nonlocal mapping
        for cc in range(c1, c2 + 1):
            name = cell_text(ws.cell(rr, cc))
            if not looks_like_name(name):
                continue
            for k in _fill_color_keys(ws.cell(rr, cc).fill):
                mapping.setdefault(k, name)

    # 1) 固定
    _add_row(56, H, AF)
    if mapping:
        return mapping

    # 2) 54-58
    for rr in range(54, 59):
        _add_row(rr, C, AZ)
    if mapping:
        return mapping

    # 3) 40-120 自動検出
    best_cnt = 0
    best_row = None
    for rr in range(40, 121):
        cnt = 0
        for cc in range(C, AZ + 1):
            name = cell_text(ws.cell(rr, cc))
            if not looks_like_name(name):
                continue
            if _fill_color_keys(ws.cell(rr, cc).fill):
                cnt += 1
        if cnt > best_cnt:
            best_cnt = cnt
            best_row = rr

    if best_row is not None and best_cnt >= 5:
        _add_row(best_row, C, AZ)

    return mapping

# ===== 講師確定（推測なし） =====
def resolve_teacher_strict(sh, legend: Dict[Tuple[str, Any], str], row: int, col: int) -> str:
    cell = merged_top_left_cell(sh, row, col)
    keys = _fill_color_keys(cell.fill)

    # 1) 完全一致（rgb/theme/indexed）
    for k in keys:
        if k in legend:
            return legend[k]

    # 2) 近傍RGB（オフが既定）
    if ALLOW_NEAR_RGB:
        rgbs = [k[1] for k in keys if k and k[0] == "rgb"]
        if rgbs:
            best = (1e9, None)
            for (lk, name) in legend.items():
                if lk[0] != "rgb":
                    continue
                for rgb in rgbs:
                    d = _rgb_dist(rgb, lk[1])
                    if d < best[0]:
                        best = (d, name)
            if best[1] and best[0] <= RGB_NEAR_THRESH:
                return best[1]

    return ""

# ===== 教室行・時間帯 =====
def detect_classroom_row(sh):
    C, AD = column_index_from_string("C"), column_index_from_string("AD")
    for r in range(1, 60):
        vals = [cell_text(sh.cell(r, c)) for c in range(C, AD + 1)]
        if sum(1 for v in vals if v and v[0] in "①②③④⑤⑥⑦⑧⑨") >= 3:
            return r
    return None

def find_timeband_above(sh, row, col, grade=None):
    """時間帯セルを上方向に探す。
    同じ列範囲に「小学部」「中学部」の2行がある場合、grade で選び分ける。
    grade: "1"-"3"=中学, "4"-"6"=小学。Noneなら従来通り最初のマッチ。
    """
    TIME_RE = re.compile(r"\d{1,2}[:：]\d{2}\s*[~～]\s*\d{1,2}[:：]\d{2}")

    # Pass 1: collect all timeband candidates from merged cells above this row
    candidates = []  # (row, full_text, time_str)
    for mr in sh.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col and mr.min_row < row:
            txt = cell_text(sh.cell(mr.min_row, mr.min_col))
            m = TIME_RE.search(txt or "")
            if m:
                candidates.append((mr.min_row, txt, m.group(0).replace("：", ":")))

    # If multiple candidates in the same column range, pick by grade label
    if len(candidates) >= 2 and grade is not None:
        is_junior = str(grade) in ("1", "2", "3")
        for r_cand, txt, time_str in candidates:
            if is_junior and "中学" in txt:
                return time_str
            if not is_junior and "小学" in txt:
                return time_str
        # fallback: no label match, return closest row above
        candidates.sort(key=lambda x: row - x[0])
        return candidates[0][2]

    if candidates:
        # Single candidate or no grade info: return first found
        candidates.sort(key=lambda x: x[0])
        return candidates[0][2]

    # Pass 2: fallback scan non-merged cells
    for r in range(row - 1, 1, -1):
        v = cell_text(sh.cell(r, col))
        if TIME_RE.search(v or ""):
            return v.replace("：", ":")
    return ""

# ===== 日付/曜日（new方式：結合セル左上＋下段救済） =====
def parse_day_week_from_row(sh, row, *, allow_prev_row: bool = False):
    def _read_on_row(rr: int):
        day = None
        w = None
        maxcol = sh.max_column
        left_date_col = 1
        right_date_col = max(1, maxcol - 1)
        left_wday_col = 2
        right_wday_col = maxcol

        for c in (left_date_col, right_date_col):
            vv = normalize_digits(merged_value_text(sh, rr, c))
            m = re.search(r"([0-9]{1,2})", vv)
            if m:
                d = int(m.group(1))
                if 1 <= d <= 31:
                    day = d
                    break

        for c in (left_wday_col, right_wday_col):
            vv = merged_value_text(sh, rr, c)
            for ch in vv:
                if ch in "月火水木金土日":
                    w = ch
                    break
            if w:
                break

        return day, w

    day, w = _read_on_row(row)
    if allow_prev_row and (day is None or w is None) and row > 1:
        d2, w2 = _read_on_row(row - 1)
        if day is None:
            day = d2
        if w is None:
            w = w2
    return day, w

# ===== クリア範囲（テンプレ仕様） =====
PROTECT_REL_CELLS = {(11, 0), (11, 1), (8, 0)}
PROTECT_REL_ROW20_RANGE = (14, 0, 5)

CLEAR_RANGES_REL: List[Tuple[int, int, int, int]] = [
    (1,  2,  8, 1),
    (2,  4,  8, 1),
    (3,  3,  8, 1),
    (4,  3,  8, 1),
    (5,  3,  8, 6),
    (11, 1,  8, 1),
    (12, 1,  8, 1),
    (13, 6,  6, 1),
    (13, 7,  8, 1),
    (14, 1,  8, 4),
]

def _is_protected(top: int, left: int, rr: int, cc: int) -> bool:
    if (rr - top, cc - left) in PROTECT_REL_CELLS:
        return True
    pr, pc0, pcw = PROTECT_REL_ROW20_RANGE
    if rr == top + pr and left + pc0 <= cc <= left + pcw:
        return True
    return False

def clear_one_block(ws, top: int, left: int):
    for drow, dcs, dce, height in CLEAR_RANGES_REL:
        for rr in range(top + drow, top + drow + height):
            for cc in range(left + dcs, left + dce + 1):
                if _is_protected(top, left, rr, cc):
                    continue
                safe_write(ws, rr, cc, "")

# ===== スロット数 =====
def count_slots_in_template(ws) -> int:
    start = 2
    step  = 10
    maxcol = ws.max_column
    return max(1, min(30, (maxcol - start) // step + 1))

# ===== ラベル＆特回数セル =====
def write_class_label(ws, top: int, left: int, label: Optional[str]):
    safe_write(ws, top + 1, left, label if label else "")

def mark_special_counters(ws, col_left: int):
    safe_write(ws, 2, col_left + 4, "特")
    safe_write(ws, 3, col_left + 5, "特")

# ===== 出力（MAIN） =====
def _day_int(ev) -> int:
    """イベントの日を整数で返す（比較用）"""
    d = str(getattr(ev, "day", "") or "")
    return int(d) if d.isdigit() else 0


def _build_merged_slots(classes: dict, order: List[str]) -> List[dict]:
    """
    特と通常を分離し、通常は授業回数でS/A/B横並び、
    特は時系列上の正しい位置に専用スロットとして挿入する。

    戻り値: [{クラス名: Event or None, "_special": bool}, ...]
    """
    specials: Dict[str, List[Event]] = {}
    regulars: Dict[str, List[Event]] = {}
    for k in order:
        specials[k] = [e for e in classes.get(k, []) if getattr(e, "special", False)]
        regulars[k] = [e for e in classes.get(k, []) if not getattr(e, "special", False)]

    # 通常イベントを授業回数でペアリング
    max_regular = max((len(regulars[k]) for k in order), default=0)
    regular_slots: List[dict] = []
    for i in range(max_regular):
        slot: dict = {"_special": False}
        for k in order:
            slot[k] = regulars[k][i] if i < len(regulars[k]) else None
        regular_slots.append(slot)

    # 特イベントの挿入位置を決定（同クラスの通常イベントの時系列位置）
    insert_items: List[Tuple[int, int, str, Event]] = []
    for k in order:
        for ev in specials[k]:
            day_val = _day_int(ev)
            insert_before = len(regulars[k])
            for ri, reg_ev in enumerate(regulars[k]):
                if day_val < _day_int(reg_ev):
                    insert_before = ri
                    break
            insert_items.append((insert_before, day_val, k, ev))
    insert_items.sort(key=lambda x: (x[0], x[1]))

    # 通常スロット列に特スロットを挿入
    merged: List[dict] = []
    sp_idx = 0
    for ri in range(max_regular):
        while sp_idx < len(insert_items) and insert_items[sp_idx][0] == ri:
            _, _, k, ev = insert_items[sp_idx]
            sp_slot: dict = {"_special": True}
            for kk in order:
                sp_slot[kk] = ev if kk == k else None
            merged.append(sp_slot)
            sp_idx += 1
        merged.append(regular_slots[ri])

    # 通常イベントの末尾以降に挿入される特（末尾の特）
    while sp_idx < len(insert_items):
        _, _, k, ev = insert_items[sp_idx]
        sp_slot = {"_special": True}
        for kk in order:
            sp_slot[kk] = ev if kk == k else None
        merged.append(sp_slot)
        sp_idx += 1

    return merged


def fill_sheet_main(ws_out, target_month: int, classes: dict, *, teacher_blank: bool = False):
    base_top = 6
    order = ["S", "A", "B"]
    idx = {k: i for i, k in enumerate(order)}
    total_slots = count_slots_in_template(ws_out)

    merged_slots = _build_merged_slots(classes, order)

    for si, slot in enumerate(merged_slots):
        if si >= total_slots:
            break
        col_left = 2 + 10 * si
        is_special = slot["_special"]

        if is_special:
            mark_special_counters(ws_out, col_left)

        for k in order:
            top = base_top + ROW_STEP * idx[k]
            clear_one_block(ws_out, top, col_left)

            ev = slot.get(k)
            if ev is None:
                label = ""
            elif is_special:
                label = "特"
            else:
                label = k
            write_class_label(ws_out, top, col_left, label)

            if ev is not None:
                safe_write(ws_out, top + 4, col_left, str(target_month) if target_month else "")
                safe_write(ws_out, top + 5, col_left, str(ev.day) if ev.day != "" else "")
                safe_write(ws_out, top + 6, col_left, ev.wday or "")
                safe_write(ws_out, top + 8, col_left, "" if teacher_blank else (ev.teacher or ""))

    # 残りの空スロットをクリア
    for si in range(len(merged_slots), total_slots):
        col_left = 2 + 10 * si
        for k in order:
            top = base_top + ROW_STEP * idx[k]
            clear_one_block(ws_out, top, col_left)
            write_class_label(ws_out, top, col_left, "")

# ===== 出力（X） =====
def fill_sheet_x(ws_out, target_month: int, x_events: List[Event], *, teacher_blank: bool = False):
    base_top = 6
    total_slots = count_slots_in_template(ws_out)
    n = len(x_events)

    for slot in range(total_slots):
        col_left = 2 + 10 * slot
        clear_one_block(ws_out, base_top, col_left)

        if slot < n:
            ev = x_events[slot]
            label = "特" if getattr(ev, "special", False) else "X"
            if getattr(ev, "special", False):
                mark_special_counters(ws_out, col_left)
        else:
            ev = None
            label = ""

        write_class_label(ws_out, base_top, col_left, label)

        if ev is not None:
            safe_write(ws_out, base_top + 4, col_left, str(target_month) if target_month else "")
            safe_write(ws_out, base_top + 5, col_left, str(ev.day) if ev.day != "" else "")
            safe_write(ws_out, base_top + 6, col_left, ev.wday or "")
            safe_write(ws_out, base_top + 8, col_left, "" if teacher_blank else (ev.teacher or ""))

# ===== 表示名 =====
def _display_subj_name(grade: str, subj: str) -> str:
    if subj == "数" and grade in ("4", "5", "6"):
        return "算数"
    return SUBJ_NAME[subj]

# ===== 回数式補修（テンプレ側） =====
def _chain_formula(prev_cells: List[str], base_cell_abs: str) -> str:
    if not prev_cells:
        return f"=IF({base_cell_abs}=\"\",\"\",{base_cell_abs})"
    prev = prev_cells[-1]
    chain = f"IF({base_cell_abs}=\"\",\"\",{base_cell_abs})"
    for cell in reversed(prev_cells):
        chain = f"IF(ISNUMBER({cell}),{cell}+1,{chain})"
    return f"=IF({prev}=\"\",\"\",IF({prev}=\"特\",{chain},{prev}+1))"

def patch_counter_formulas(ws):
    total_slots = count_slots_in_template(ws)
    annual_cells = []
    week_cells = []
    for slot in range(total_slots):
        col_left = 2 + 10 * slot
        annual_cells.append(ws.cell(row=2, column=col_left + 4).coordinate)
        week_cells.append(ws.cell(row=3, column=col_left + 5).coordinate)

    for i in range(1, len(annual_cells)):
        ws[annual_cells[i]].value = _chain_formula(annual_cells[:i], "$FG$2")
    for i in range(1, len(week_cells)):
        ws[week_cells[i]].value = _chain_formula(week_cells[:i], "$FG$3")

# ===== シートコピー安全版（StyleProxy対策） =====
def copy_worksheet_contents_safe(src, dst):
    for col, dim in src.column_dimensions.items():
        dd = dst.column_dimensions[col]
        dd.width = dim.width
        dd.hidden = dim.hidden
        dd.outlineLevel = dim.outlineLevel
        dd.collapsed = dim.collapsed

    for row, dim in src.row_dimensions.items():
        rd = dst.row_dimensions[row]
        rd.height = dim.height
        rd.hidden = dim.hidden
        rd.outlineLevel = dim.outlineLevel
        rd.collapsed = dim.collapsed

    try:
        dst.freeze_panes = src.freeze_panes
    except Exception:
        pass

    for merged in list(src.merged_cells.ranges):
        dst.merge_cells(str(merged))

    # データ入力規則（プルダウン等）をコピー
    if src.data_validations and src.data_validations.dataValidation:
        for dv in src.data_validations.dataValidation:
            dst.add_data_validation(copy(dv))

    max_row = src.max_row
    max_col = src.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            s_cell = src.cell(r, c)
            if isinstance(s_cell, MergedCell):
                continue
            d_cell = dst.cell(r, c)
            d_cell.value = s_cell.value

            if s_cell.has_style:
                d_cell.font = copy(s_cell.font)
                d_cell.fill = copy(s_cell.fill)
                d_cell.border = copy(s_cell.border)
                d_cell.alignment = copy(s_cell.alignment)
                d_cell.number_format = s_cell.number_format
                d_cell.protection = copy(s_cell.protection)
            else:
                d_cell.number_format = s_cell.number_format

# ===== 破損xlsx処理（厳格） =====
def quarantine_broken_xlsx(path: Path, reason: str) -> None:
    for i in range(1, 100):
        q = path.with_name(f"{path.stem}__broken{i}{path.suffix}")
        if q.exists():
            continue
        try:
            path.rename(q)
            print(f"[WARN] 破損xlsxを退避しました: {q.name}（理由: {reason}）")
            return
        except Exception:
            break
    # 退避できない＝ロック中
    raise RuntimeError(
        f"破損xlsxを退避できません（ファイルがロックされています）: {path.name}\n"
        f"Excelで開いていないか確認し、手動で削除/リネームしてください。"
    )

def load_workbook_safe(path: Path) -> Optional[openpyxl.Workbook]:
    try:
        return openpyxl.load_workbook(path)
    except (zipfile.BadZipFile, KeyError, OSError, ValueError) as e:
        quarantine_broken_xlsx(path, f"{type(e).__name__}: {e}")
        return None

# ===== 年ファイル運用 =====
def open_or_create_year_workbook(out_path: Path, template_path: Path, hidden_name: str) -> openpyxl.Workbook:
    if out_path.exists():
        wb = load_workbook_safe(out_path)
        if wb is not None:
            return wb
        # ここには来ない（退避失敗なら例外で停止）
    wb = openpyxl.load_workbook(template_path)
    ws0 = wb[wb.sheetnames[0]]
    for name in list(wb.sheetnames)[1:]:
        del wb[name]
    ws0.title = hidden_name
    ws0.sheet_state = "hidden"
    patch_counter_formulas(ws0)
    return wb

def ensure_hidden_template_sheet(wb: openpyxl.Workbook, template_path: Path, hidden_name: str):
    if hidden_name in wb.sheetnames:
        ws = wb[hidden_name]
        ws.sheet_state = "hidden"
        patch_counter_formulas(ws)
        return ws

    t_wb = openpyxl.load_workbook(template_path)
    src = t_wb[t_wb.sheetnames[0]]
    ws = wb.create_sheet(hidden_name)
    copy_worksheet_contents_safe(src, ws)
    ws.sheet_state = "hidden"
    patch_counter_formulas(ws)
    return ws

def create_month_sheet(wb: openpyxl.Workbook, hidden_ws, year: int, month: int):
    base = month_sheet_base_name(year, month)

    # 既に同じ月シートがある場合は何もしない
    if base in wb.sheetnames:
        return None

    # hidden template から複製
    ws = wb.copy_worksheet(hidden_ws)
    ws.title = base
    ws.sheet_state = "visible"

    # データ入力規則が copy_worksheet で欠落する場合の補完
    if hidden_ws.data_validations and hidden_ws.data_validations.dataValidation:
        existing = {str(dv.sqref) for dv in (ws.data_validations.dataValidation or [])}
        for dv in hidden_ws.data_validations.dataValidation:
            if str(dv.sqref) not in existing:
                ws.add_data_validation(copy(dv))

    # 右端へ移動
    try:
        current_index = wb.worksheets.index(ws)
        rightmost_index = len(wb.worksheets) - 1
        offset = rightmost_index - current_index
        if offset != 0:
            wb.move_sheet(ws, offset=offset)
    except Exception:
        pass

    try:
        wb.active = wb.sheetnames.index(base)
    except Exception:
        pass
    return ws

def set_header_cells(ws, campus: str, grade_label: str, subject_name: str, month: int):
    def set_if_not_formula(addr: str, value: Any):
        v = ws[addr].value
        if isinstance(v, str) and v.startswith("="):
            return
        ws[addr].value = value

    set_if_not_formula("B2", campus)
    set_if_not_formula("I2", grade_label)
    set_if_not_formula("I3", subject_name)
    set_if_not_formula("E3", month)

    f2 = ws["F2"].value
    g3 = ws["G3"].value
    ws["FG2"].value = f2 if isinstance(f2, (int, float)) else ""
    ws["FG3"].value = g3 if isinstance(g3, (int, float)) else ""

def save_year_workbook(wb: openpyxl.Workbook, out_path: Path):
    if all(wb[sn].sheet_state != "visible" for sn in wb.sheetnames):
        wb.create_sheet("INIT").sheet_state = "visible"
    wb.save(out_path)

# ===== イベント収集（重複除去＋未判定講師ログ） =====
def collect_events(sh, target_month: int, campus: str, year: int, month: int) -> List[Event]:
    legend = detect_teacher_legend(sh)
    room_row = detect_classroom_row(sh)
    C, AD = column_index_from_string("C"), column_index_from_string("AD")

    evs = []
    seen = set()
    undecided = []

    for r in range(2, 55):
        day, w = parse_day_week_from_row(sh, r)
        for c in range(C, AD + 1):
            txt = merged_value_text(sh, r, c)
            if not txt:
                continue
            tok = parse_class_token(txt)
            if not tok:
                continue

            if day is None or w is None:
                day, w = parse_day_week_from_row(sh, r, allow_prev_row=True)

            g, k, subj, special = tok

            teacher = resolve_teacher_strict(sh, legend, r, c)
            if teacher == "":
                # 「講師が決まらない」セルを記録（間違えないため空欄）
                coord = f"{get_column_letter(c)}{r}"
                fill_keys = list(_fill_color_keys(merged_top_left_cell(sh, r, c).fill))
                undecided.append((coord, txt, fill_keys))

            time = find_timeband_above(sh, r, c, grade=g)
            classroom = merged_value_text(sh, room_row, c) if room_row else ""

            ev = Event(target_month, day or "", w or "", time, classroom, g, k, subj, txt, teacher, special, r, c)

            key = (ev.day, ev.wday, ev.time, ev.classroom, ev.grade, ev.klass, ev.subj, ev.text, ev.teacher, ev.special)
            if key in seen:
                continue
            seen.add(key)
            evs.append(ev)

    if undecided:
        fname = f"未判定講師_{campus}_{year:04d}-{month:02d}.txt"
        with open(fname, "w", encoding="utf-8") as f:
            f.write(f"未判定講師一覧（{campus} {year:04d}-{month:02d}）\n")
            f.write("講師色が凡例と一致しない/読み取れないセルです。講師は空欄にしています（誤判定防止）。\n\n")
            for coord, txt, keys in undecided:
                f.write(f"- {coord}  {txt}\n")
                f.write(f"    fill_keys: {keys}\n")
        print(f"[WARN] 講師未判定セルあり: {fname} を確認してください。")

    evs.sort(key=lambda e: (int(e.day) if str(e.day).isdigit() else 99, e.r, e.c))
    return evs

# ===== 重複割当チェック（講師が空欄のものは対象外） =====
def collision_check(events: List[Event], campus: str, year: int, month: int):
    rows = []
    by_key = defaultdict(list)
    for e in events:
        if not e.teacher:
            continue
        key = (e.day, e.time, e.teacher)
        by_key[key].append(e)

    for (day, time, teacher), lst in sorted(by_key.items(), key=lambda x: (int(x[0][0]) if str(x[0][0]).isdigit() else 99, x[0][1], x[0][2])):
        rooms = sorted(set([x.classroom for x in lst if x.classroom]))
        if len(rooms) >= 2 or len(lst) >= 2:
            rows.append((day, time, teacher, rooms, [(x.grade, x.klass, x.subj, x.text) for x in lst]))

    if not rows:
        return

    fname = f"重複割当チェック_{campus}_{year:04d}-{month:02d}.txt"
    with open(fname, "w", encoding="utf-8") as f:
        f.write(f"重複割当チェック（{campus} {year:04d}-{month:02d}）\n")
        f.write("同一日・同一時間帯で同一講師が複数教室に割り当てられていないか確認してください。\n\n")
        for day, time, teacher, rooms, items in rows:
            f.write(f"- {day}日 {time}  講師:{teacher}  教室:{', '.join(rooms) if rooms else '(不明)'}\n")
            for g, k, subj, text in items:
                f.write(f"    ・{g}{k}{subj}  {text}\n")
            f.write("\n")
    print(f"[WARN] 重複割当の可能性: {fname} を確認してください。")

# ===== 校舎単位の出力 =====
def export_one_campus(ws_schedule, campus: str, year: int, month: int):
    all_events = collect_events(ws_schedule, month, campus, year, month)
    collision_check(all_events, campus, year, month)

    # 本体（学年×科目）は補講を除外
    buckets = defaultdict(list)
    for e in all_events:
        if ("英語補講" in e.text) or ("数学補講" in e.text):
            continue
        buckets[(e.grade, e.subj, e.klass)].append(e)

    for grade in GRADES:
        for subj in SUBJS:
            s_list = buckets.get((grade, subj, "S"), [])
            a_list = buckets.get((grade, subj, "A"), [])
            b_list = buckets.get((grade, subj, "B"), [])
            x_list = buckets.get((grade, subj, "X"), [])

            subj_j = _display_subj_name(grade, subj)
            grade_j = GRADE_LABEL[grade]

            if any([s_list, a_list, b_list]):
                out_name = f"{campus}{grade_j}{subj_j}_{year}.xlsx"
                out_path = Path(out_name)

                wb = open_or_create_year_workbook(out_path, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")
                tmpl = ensure_hidden_template_sheet(wb, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")

                ws_out = create_month_sheet(wb, tmpl, year, month)
                if ws_out is None:
                    print(f"[SKIP] {out_path.name}: {year:04d}-{month:02d} は既にあります")
                else:
                    set_header_cells(ws_out, campus, grade_j, subj_j, month)
                    fill_sheet_main(ws_out, month, {"S": s_list, "A": a_list, "B": b_list}, teacher_blank=False)
                    save_year_workbook(wb, out_path)
                    print(f"[OK] {out_path.name} + {ws_out.title}")

            if x_list:
                out_name = f"{campus}{grade_j}X{subj_j}_{year}.xlsx"
                out_path = Path(out_name)

                wb = open_or_create_year_workbook(out_path, TEMPLATE_X, "__TEMPLATE_X__")
                tmpl = ensure_hidden_template_sheet(wb, TEMPLATE_X, "__TEMPLATE_X__")

                ws_out = create_month_sheet(wb, tmpl, year, month)
                if ws_out is None:
                    print(f"[SKIP] {out_path.name}: {year:04d}-{month:02d} は既にあります")
                else:
                    set_header_cells(ws_out, campus, grade_j, subj_j, month)
                    fill_sheet_x(ws_out, month, x_list, teacher_blank=False)
                    save_year_workbook(wb, out_path)
                    print(f"[OK] {out_path.name} + {ws_out.title}")

    # 補講ファイル
    for keyword, title in [("英語補講", "英語補講"), ("数学補講", "数学補講")]:
        hits = [e for e in all_events if keyword in e.text]
        if not hits:
            continue

        out_name = f"{campus}{title}_{year}.xlsx"
        out_path = Path(out_name)

        wb = open_or_create_year_workbook(out_path, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")
        tmpl = ensure_hidden_template_sheet(wb, TEMPLATE_MAIN, "__TEMPLATE_MAIN__")

        ws_out = create_month_sheet(wb, tmpl, year, month)
        if ws_out is None:
            print(f"[SKIP] {out_path.name}: {year:04d}-{month:02d} は既にあります")
        else:
            set_header_cells(ws_out, campus, "", title, month)
            fill_sheet_main(ws_out, month, {"S": hits, "A": [], "B": []}, teacher_blank=True)
            save_year_workbook(wb, out_path)
            print(f"[OK] {out_path.name} + {ws_out.title}")

# ===== シート選別（教務部用） =====
def _norm(s: str) -> str:
    return (s or "").replace(" ", "").replace("\u3000", "")

def choose_target_sheets(wb):
    picks = []
    for name in wb.sheetnames:
        nm = _norm(name)
        if "教務" not in nm:
            continue
        if "本校" in nm:
            picks.append(("本校", name, wb[name]))
        elif "南教室" in nm or ("南" in nm):
            picks.append(("南教室", name, wb[name]))

    result = []
    seen = set()
    for campus, name, sh in picks:
        if campus in seen:
            continue
        result.append((campus, name, sh))
        seen.add(campus)
    return result

# ===== メイン =====
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--prefer", type=str, help="優先月カンマ区切り（例: 11,10）")
    args = ap.parse_args()
    prefer = [int(x) for x in re.split(r"[,\s]+", args.prefer.strip()) if x] if args.prefer else None

    if not TEMPLATE_MAIN.exists() or not TEMPLATE_X.exists():
        raise FileNotFoundError("TEMPLATE_MAIN.xlsx / TEMPLATE_X.xlsx が見つかりません（同フォルダに置いてください）。")

    y, m, sch = pick_schedule_in_same_folder(prefer)
    print(f"[INFO] スケジュール: {sch.name} ({y}-{m:02d})")
    wb_s = openpyxl.load_workbook(sch, data_only=True, keep_vba=True)

    targets = choose_target_sheets(wb_s)
    if not targets:
        print("対象シート（本校/南教室の教務部用）が見つかりません。")
        return

    for campus, sname, sh in targets:
        print(f"[INFO] 対象: {campus} / {sname}")
        export_one_campus(sh, campus, y, m)

if __name__ == "__main__":
    main()