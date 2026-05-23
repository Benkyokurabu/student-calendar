#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
プログラム③：報告抽出

役割
- 授業日誌Excelから、Web掲載用の本文だけを抽出して JSON にする
- 読み取る項目
  - D7   授業内容       -> content
  - F8   ページ         -> page
  - E9:E10 宿題        -> homework
  - E11  記録           -> report
  - D17  録画URL        -> recordingUrl
  - B14  担当講師       -> teacher
  - C18:C19 欠席       -> absence (2行を改行連結)
  - I20  カリキュラム進捗 -> curriculumProgress
  - ヘッダー: 第○回 / ○月○週  -> sessionNumber / monthNum / weekNum

出力
- journal_latest.json
- journal_YYYY-MM.json

前提
- カレ ンダー側 repo に schedule_latest.json があること
- 可能なら journal_map_latest.json もあること
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

TARGET_FOLDER_NAME = "09　授業日誌"
MONTH_SHEET_RE = re.compile(r"^(\d{4})-(\d{2})(?:\(再(\d+)\))?$")

CAMPUS_JP = {"hon": "本校", "minami": "南教室"}
GRADE_JP = {
    "e4": "小４",
    "e5": "小５",
    "e6": "小６",
    "j1": "中１",
    "j2": "中２",
    "j3": "中３",
}
SUBJECT_JP = {
    "eng": "英語",
    "math": "数学",
    "jp": "国語",
    "sci": "理科",
    "soc": "社会",
    "arith": "算数",
}

FIRST_BLOCK_COL = 2
BLOCK_WIDTH = 10
ROW_STEP = 20
BASE_TOP_MAIN = {"S": 6, "A": 26, "B": 46}
BASE_TOP_X = 6


def get_default_repo_dir() -> Optional[Path]:
    user_home = Path.home()
    candidates = [
        user_home / "OneDrive" / "デスクトップ" / "生徒スケジュール表",
        Path(r"C:\Users\kudok\OneDrive\デスクトップ\生徒スケジュール表"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def get_default_journal_dir() -> Optional[Path]:
    user_home = Path.home()
    candidates = [
        user_home / "OneDrive" / "●勉強クラブ共有" / TARGET_FOLDER_NAME,
        Path(r"C:\Users\kudok\OneDrive\●勉強クラブ共有\09　授業日誌"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def merged_top_left_cell(ws, row: int, col: int):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col)
    return ws.cell(row, col)


def read_merged_text(ws, row: int, col: int) -> str:
    v = merged_top_left_cell(ws, row, col).value
    return "" if v is None else str(v).strip()


def safe_write_json(path: Path, data: dict) -> None:
    """一時ファイルに書いてからリネームする安全な書き込み"""
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def normalize_time(s: str) -> str:
    return str(s or "").replace("~", "～").strip()


def make_entry_key(ev: dict) -> str:
    return "|".join(
        [
            str(ev.get("date", "")),
            normalize_time(ev.get("time", "")),
            str(ev.get("campus", "")),
            str(ev.get("groupKey", "")),
            str(ev.get("room", "")),
        ]
    )


SKIP_DIRS = {"_backup", "退避", "__pycache__"}


def find_workbook(journal_dir: Path, filename: str) -> Optional[Path]:
    p = journal_dir / filename
    if p.exists():
        return p
    for candidate in journal_dir.rglob(filename):
        if candidate.is_file():
            # _backup / 退避 など除外
            if SKIP_DIRS & {part for part in candidate.relative_to(journal_dir).parts}:
                continue
            return candidate
    return None


def find_month_sheet(wb: openpyxl.Workbook, year: int, month: int):
    target = f"{year:04d}-{month:02d}"
    if target in wb.sheetnames:
        return wb[target]

    cands: List[Tuple[int, str]] = []
    for name in wb.sheetnames:
        m = MONTH_SHEET_RE.match(name)
        if not m:
            continue
        y = int(m.group(1))
        mo = int(m.group(2))
        retry = int(m.group(3) or 0)
        if y == year and mo == month:
            cands.append((retry, name))

    if cands:
        cands.sort()
        return wb[cands[-1][1]]
    return None


def _day_from_date(ev: dict) -> int:
    """イベントの日付文字列から日を整数で返す（比較用）"""
    d = str(ev.get("date", ""))
    if len(d) >= 10:
        dd = d[8:10]
        return int(dd) if dd.isdigit() else 0
    return 0


def _build_merged_slots_for_extract(classes: Dict[str, List[dict]], order: List[str]) -> List[dict]:
    """export_by_grade_subject._build_merged_slots と同じロジック。
    特スロットを正しい位置に挿入し、Excelのカラム配置を再現する。"""
    specials: Dict[str, List[dict]] = {}
    regulars: Dict[str, List[dict]] = {}
    for k in order:
        specials[k] = [e for e in classes.get(k, []) if e.get("special", False)]
        regulars[k] = [e for e in classes.get(k, []) if not e.get("special", False)]

    max_regular = max((len(regulars[k]) for k in order), default=0)

    # 特イベントの挿入位置を決定
    insert_items: List[Tuple[int, int, str, dict]] = []
    for k in order:
        for ev in specials[k]:
            day_val = _day_from_date(ev)
            insert_before = len(regulars[k])
            for ri, reg_ev in enumerate(regulars[k]):
                if day_val < _day_from_date(reg_ev):
                    insert_before = ri
                    break
            insert_items.append((insert_before, day_val, k, ev))
    insert_items.sort(key=lambda x: (x[0], x[1]))

    # 通常スロット列に特スロットを挿入
    merged: List[dict] = []
    sp_idx = 0
    for ri in range(max_regular):
        while sp_idx < len(insert_items) and insert_items[sp_idx][0] == ri:
            _, _, k, ev = insert_items[sp_idx]
            sp_slot: dict = {"_special": True}
            for kk in order:
                sp_slot[kk] = ev if kk == k else None
            merged.append(sp_slot)
            sp_idx += 1
        reg_slot: dict = {"_special": False}
        for k in order:
            reg_slot[k] = regulars[k][ri] if ri < len(regulars[k]) else None
        merged.append(reg_slot)

    while sp_idx < len(insert_items):
        _, _, k, ev = insert_items[sp_idx]
        sp_slot = {"_special": True}
        for kk in order:
            sp_slot[kk] = ev if kk == k else None
        merged.append(sp_slot)
        sp_idx += 1

    return merged


def build_group_slot_map(events: List[dict], slots_map: Optional[dict]) -> Dict[str, int]:
    result: Dict[str, int] = {}

    # Excelカラム位置を再現するため、_build_merged_slots と同じロジックで
    # S/A/B 全クラスをまとめてマージスロットを構築する。
    # ※ journal_map の per-groupKey インデックスは、他クラスの特スロット挿入を
    #   考慮できずズレるため使用しない。
    # (campus, grade, subject) = 1つのExcelワークブックに対応
    wb_groups: Dict[Tuple[str, str, str], Dict[str, List[dict]]] = defaultdict(lambda: defaultdict(list))
    for ev in events:
        campus = str(ev.get("campus", ""))
        grade = str(ev.get("grade", ""))
        subject = str(ev.get("subject", ""))
        klass = str(ev.get("class", "") or "")
        if not campus or not grade or not subject or not klass:
            continue
        wb_groups[(campus, grade, subject)][klass].append(ev)

    for wb_key, classes in wb_groups.items():
        if "X" in classes:
            # X クラス: 単純な連番（fill_sheet_x と同じ）
            for i, ev in enumerate(classes["X"]):
                key = make_entry_key(ev)
                if key not in result:
                    result[key] = i + 1

        # S/A/B: 特スロット挿入を含むマージ済みスロットで位置を決定
        order = ["S", "A", "B"]
        if not any(classes.get(k) for k in order):
            continue
        merged = _build_merged_slots_for_extract(classes, order)
        for si, slot in enumerate(merged):
            for k in order:
                ev = slot.get(k)
                if ev is not None:
                    key = make_entry_key(ev)
                    if key not in result:
                        result[key] = si + 1

    return result


def subject_for_filename(grade_key: str, subj_key: str) -> str:
    """小学生(e4,e5,e6)の math は '算数' に変換"""
    if subj_key == "math" and grade_key in ("e4", "e5", "e6"):
        return "算数"
    return SUBJECT_JP.get(subj_key, "")


def workbook_filename_from_event(ev: dict, year: int) -> Optional[str]:
    campus = CAMPUS_JP.get(str(ev.get("campus", "")))
    grade = GRADE_JP.get(str(ev.get("grade", "")))
    grade_key = str(ev.get("grade", ""))
    subj_key = str(ev.get("subject", ""))
    subject = subject_for_filename(grade_key, subj_key)
    klass = str(ev.get("class", "") or "")
    if not campus or not grade or not subject:
        return None

    if klass == "X":
        return f"{campus}{grade}X{subject}_{year}.xlsx"
    return f"{campus}{grade}{subject}_{year}.xlsx"


def read_block(ws, top_row: int, left_col: int) -> dict:
    content = read_merged_text(ws, top_row + 1, left_col + 2)       # D7
    page = read_merged_text(ws, top_row + 2, left_col + 4)          # F8
    hw1 = read_merged_text(ws, top_row + 3, left_col + 3)           # E9
    hw2 = read_merged_text(ws, top_row + 4, left_col + 3)           # E10
    report = read_merged_text(ws, top_row + 5, left_col + 3)        # E11
    recording_url = read_merged_text(ws, top_row + 11, left_col + 2)  # D17
    teacher = read_merged_text(ws, top_row + 8, left_col)            # B14 (担当)
    abs1 = read_merged_text(ws, top_row + 12, left_col + 1)         # C18 (欠席1行目)
    abs2 = read_merged_text(ws, top_row + 13, left_col + 1)         # C19 (欠席2行目)
    absence = "\n".join([x for x in [abs1, abs2] if x])
    curriculum_sign = read_merged_text(ws, top_row + 14, left_col + 6)  # H20 (進捗符号: +/-/±)
    curriculum_val = read_merged_text(ws, top_row + 14, left_col + 7)   # I20 (進捗数値)
    curriculum = (curriculum_sign + curriculum_val) if (curriculum_sign or curriculum_val) else ""
    note = read_merged_text(ws, top_row + 15, left_col + 1)          # C21 (備考)

    homework = [x for x in [hw1, hw2] if x]
    return {
        "content": content,
        "page": page,
        "homework": homework,
        "report": report,
        "recordingUrl": recording_url,
        "teacher": teacher,
        "absence": absence,
        "curriculumProgress": curriculum,
        "note": note,
    }


class WorkbookCache:
    """同じExcelを何度も開かないようにキャッシュする"""
    def __init__(self, journal_dir: Path, temp_dir: Path):
        self.journal_dir = journal_dir
        self.temp_dir = temp_dir
        self._wb_cache: Dict[str, Optional[openpyxl.Workbook]] = {}
        self._path_cache: Dict[str, Optional[Path]] = {}

    def get_workbook(self, filename: str) -> Optional[openpyxl.Workbook]:
        if filename in self._wb_cache:
            return self._wb_cache[filename]

        wb_path = find_workbook(self.journal_dir, filename)
        if wb_path is None:
            self._wb_cache[filename] = None
            return None

        try:
            tmp_file = self.temp_dir / filename
            if not tmp_file.exists():
                shutil.copy2(wb_path, tmp_file)
            wb = openpyxl.load_workbook(tmp_file, data_only=True, read_only=False)
            self._wb_cache[filename] = wb
            return wb
        except Exception:
            self._wb_cache[filename] = None
            return None

    def close_all(self):
        for wb in self._wb_cache.values():
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


_DAY_ROWS = [11, 31, 51]  # S(row 11), A(row 31), B(row 51) の日付行


def _slot_has_day(ws, col_left: int) -> bool:
    """スロットに日付があるか（S/A/Bいずれかの日付行をチェック）"""
    for row in _DAY_ROWS:
        v = ws.cell(row=row, column=col_left).value
        if v is not None and str(v).strip() != "":
            return True
    return False


def _compute_annual_start_from_wb(wb, year: int, month: int) -> int | None:
    """前月シートの最終回 + 1 を計算する（export_by_grade_subject.compute_annual_start と同等）"""
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    prev_name = f"{prev_year:04d}-{prev_month:02d}"
    if prev_name not in wb.sheetnames:
        return 1
    prev_ws = wb[prev_name]
    # 前月のF2から最終回を計算
    prev_f2 = prev_ws["F2"].value
    if isinstance(prev_f2, str) and prev_f2.strip() == "特":
        prev_f2 = prev_ws.cell(row=2, column=163).value
    if not isinstance(prev_f2, (int, float)):
        return _compute_annual_start_from_wb(wb, prev_year, prev_month)
    current = int(prev_f2)
    found_first = False
    for slot in range(17):
        col = FIRST_BLOCK_COL + slot * BLOCK_WIDTH
        if not _slot_has_day(prev_ws, col):
            continue
        ann = prev_ws.cell(row=2, column=col + 4).value
        if isinstance(ann, str) and ann.strip() == "特":
            continue
        if not found_first:
            found_first = True
        else:
            current += 1
    return (current + 1) if found_first else 1


def _compute_slot_session(ws, target_left_col: int, *, wb=None, year: int = 0, month: int = 0) -> tuple:
    """F2の値とスロット位置から年回数・月回数を計算する。
    Excelの数式と同じロジック: F2から開始し、通常スロットごとに+1、「特」はスキップ。
    Returns (sessionNumber, monthNum, weekNum)
    """
    f2 = ws["F2"].value
    if isinstance(f2, str) and f2.strip() == "特":
        # FG2 (col 163) にバックアップ値がある
        fg2 = ws.cell(row=2, column=163).value
        # FG2がテンプレートのデフォルト値の可能性があるため、前月から再計算して検証
        if wb is not None and year and month:
            computed = _compute_annual_start_from_wb(wb, year, month)
            if computed is not None:
                f2 = computed
            elif isinstance(fg2, (int, float)):
                f2 = fg2
        elif isinstance(fg2, (int, float)):
            f2 = fg2
    if not isinstance(f2, (int, float)):
        return ("", "", "")

    # E3 = 月番号, G3 = 月内カウンタ開始(常に1)
    month_n_raw = ws.cell(row=3, column=5).value  # E3
    if month_n_raw is None:
        return ("", "", "")
    month_n = str(int(month_n_raw)) if isinstance(month_n_raw, (int, float)) else str(month_n_raw).strip()

    current_annual = int(f2)
    current_monthly = 1
    sheet_month = int(month_n_raw) if isinstance(month_n_raw, (int, float)) else None

    for slot in range(17):  # 最大17スロット
        col_left = FIRST_BLOCK_COL + slot * BLOCK_WIDTH
        if not _slot_has_day(ws, col_left):
            continue
        # 「特」スロットかチェック (row 2, col_left + 4)
        annual_val = ws.cell(row=2, column=col_left + 4).value
        if isinstance(annual_val, str) and annual_val.strip() == "特":
            if col_left == target_left_col:
                return ("特", "", "特")
            continue
        # 月番号がシートの月と異なるスロットは除外
        slot_month_val = ws.cell(row=3, column=col_left + 3).value
        if slot_month_val is not None:
            try:
                if int(slot_month_val) != sheet_month:
                    continue
            except (ValueError, TypeError):
                pass

        if col_left == target_left_col:
            return (str(current_annual), month_n, str(current_monthly))

        current_annual += 1
        current_monthly += 1

    return ("", "", "")


def read_slot_header(ws, left_col: int, *, wb=None, year: int = 0, month: int = 0) -> dict:
    """スロットヘッダー情報を読み取る（第○回、○月○週）
    セルに値があればそれを使い、数式未キャッシュ(None)の場合はF2から計算する。
    """
    session = read_merged_text(ws, 2, left_col + 4)   # F2 相当
    month_n = read_merged_text(ws, 3, left_col + 3)    # E3 相当
    week_n = read_merged_text(ws, 3, left_col + 5)     # G3 相当

    # F2が「特」の場合、FG2のキャッシュ値が不正な可能性があるため常に再計算
    f2_raw = ws["F2"].value
    force_recalc = isinstance(f2_raw, str) and f2_raw.strip() == "特" and wb is not None

    # 数式未キャッシュで空の項目があるか、F2=特で再計算が必要な場合
    if not session or not month_n or not week_n or force_recalc:
        calc_s, calc_m, calc_w = _compute_slot_session(ws, left_col, wb=wb, year=year, month=month)
        if not session or force_recalc:
            session = calc_s or session
        if not month_n or force_recalc:
            month_n = calc_m or month_n
        if not week_n or force_recalc:
            week_n = calc_w or week_n

    return {
        "sessionNumber": session,
        "monthNum": month_n,
        "weekNum": week_n,
    }


def _block_has_data(block: dict) -> bool:
    """ブロックに実質的なデータがあるか"""
    return bool(block.get("content") or block.get("page") or
                block.get("report") or block.get("homework"))


def _find_last_slot_col(ws) -> int | None:
    """シート内の最後のスロット（日付あり）の left_col を返す"""
    last_col = None
    for slot in range(17):
        col = FIRST_BLOCK_COL + slot * BLOCK_WIDTH
        if _slot_has_day(ws, col):
            last_col = col
    return last_col


def _read_slot_date(ws, top_row: int, left_col: int, year: int) -> str:
    """スロットの日付を YYYY-MM-DD 形式で読み取る"""
    m_val = read_merged_text(ws, top_row + 4, left_col)
    d_val = read_merged_text(ws, top_row + 5, left_col)
    if not m_val or not d_val:
        return ""
    try:
        return f"{year}-{int(m_val):02d}-{int(d_val):02d}"
    except (ValueError, TypeError):
        return ""


def _read_prev_entry(wb, ws, year: int, month: int, slot_index: int,
                     klass: str, top_row: int) -> dict | None:
    """前回スロットのデータをExcelから直接読み取る。
    同月シート内で前のスロットを遡り、なければ前月シートの最後のスロットを探す。
    """
    current_left_col = FIRST_BLOCK_COL + (max(1, slot_index) - 1) * BLOCK_WIDTH

    # --- 同月内で前のスロットを探す ---
    col = current_left_col - BLOCK_WIDTH
    while col >= FIRST_BLOCK_COL:
        if _slot_has_day(ws, col):
            block = read_block(ws, top_row, col)
            if _block_has_data(block):
                block.update(read_slot_header(ws, col, wb=wb, year=year, month=month))
                block["date"] = _read_slot_date(ws, top_row, col, year)
                return block
        col -= BLOCK_WIDTH

    # --- 前月シートを探す ---
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    prev_ws = find_month_sheet(wb, prev_year, prev_month)
    if prev_ws is None:
        return None

    last_col = _find_last_slot_col(prev_ws)
    if last_col is None:
        return None

    col = last_col
    while col >= FIRST_BLOCK_COL:
        if _slot_has_day(prev_ws, col):
            block = read_block(prev_ws, top_row, col)
            if _block_has_data(block):
                block.update(read_slot_header(prev_ws, col, wb=wb, year=prev_year, month=prev_month))
                block["date"] = _read_slot_date(prev_ws, top_row, col, prev_year)
                return block
        col -= BLOCK_WIDTH

    return None


def extract_entry_for_event(ev: dict, slot_index: int, wb_cache: WorkbookCache) -> dict:
    empty = {"content": "", "page": "", "homework": [], "report": "", "recordingUrl": "",
             "teacher": "", "absence": "", "curriculumProgress": "", "note": "",
             "sessionNumber": "", "monthNum": "", "weekNum": ""}
    date = str(ev.get("date", ""))
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", date)
    if not m:
        return empty

    year = int(m.group(1))
    month = int(m.group(2))
    klass = str(ev.get("class", "") or "")

    filename = workbook_filename_from_event(ev, year)
    if not filename:
        return empty

    wb = wb_cache.get_workbook(filename)
    if wb is None:
        return empty

    ws = find_month_sheet(wb, year, month)
    if ws is None:
        return empty

    left_col = FIRST_BLOCK_COL + (max(1, slot_index) - 1) * BLOCK_WIDTH
    if klass == "X":
        top_row = BASE_TOP_X
    else:
        top_row = BASE_TOP_MAIN.get(klass)
        if top_row is None:
            return empty

    block = read_block(ws, top_row, left_col)
    block.update(read_slot_header(ws, left_col, wb=wb, year=year, month=month))

    # 前回データをExcelから直接読み取り
    prev = _read_prev_entry(wb, ws, year, month, slot_index, klass, top_row)
    if prev:
        block["prevEntry"] = prev

    return block


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def determine_month(events: List[dict], month_arg: Optional[str]) -> str:
    if month_arg:
        if not re.match(r"^\d{4}-\d{2}$", month_arg):
            raise ValueError("--month は YYYY-MM 形式で指定してください。")
        return month_arg

    months = sorted({str(ev.get("date", ""))[:7] for ev in events if str(ev.get("date", ""))[:7]})
    if not months:
        raise ValueError("schedule_latest.json から年月を判定できませんでした。")
    return months[-1]


def find_schedule_json(repo_dir: Path, month_arg: Optional[str]) -> Tuple[Path, str]:
    """スケジュールJSONを探す。--month 指定時は月別ファイルを優先。"""
    if month_arg:
        if not re.match(r"^\d{4}-\d{2}$", month_arg):
            raise ValueError("--month は YYYY-MM 形式で指定してください。")
        # 月別ファイルを優先
        month_path = repo_dir / f"schedule_{month_arg}.json"
        if month_path.exists():
            events = load_json(month_path)
            print(f"[INFO] スケジュール: {month_path.name}")
            return month_path, month_arg
        # なければ latest から該当月をフィルタ
        latest = repo_dir / "schedule_latest.json"
        if latest.exists():
            events = load_json(latest)
            has_month = any(str(ev.get("date", ""))[:7] == month_arg for ev in events)
            if has_month:
                print(f"[INFO] スケジュール: {latest.name} (月フィルタ: {month_arg})")
                return latest, month_arg
        raise FileNotFoundError(
            f"schedule_{month_arg}.json も schedule_latest.json({month_arg}分) も見つかりません: {repo_dir}"
        )

    # --month 未指定 → latest から自動判定
    latest = repo_dir / "schedule_latest.json"
    if not latest.exists():
        raise FileNotFoundError(f"schedule_latest.json が見つかりません: {repo_dir}")
    events = load_json(latest)
    target_month = determine_month(events, None)
    print(f"[INFO] スケジュール: {latest.name} (自動判定: {target_month})")
    return latest, target_month


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo-dir", help="student-calendar リポジトリの場所。未指定時は自動探索")
    ap.add_argument("--journal-dir", help="授業日誌フォルダ。未指定時は OneDrive 共有フォルダを自動探索")
    ap.add_argument("--month", help="出力対象月（YYYY-MM）。未指定時は schedule_latest.json から自動判定")
    args = ap.parse_args()

    repo_dir = Path(args.repo_dir).resolve() if args.repo_dir else get_default_repo_dir()
    if repo_dir is None or not repo_dir.exists():
        raise FileNotFoundError("student-calendar リポジトリが見つかりません。--repo-dir で指定してください。")
    journal_dir = Path(args.journal_dir) if args.journal_dir else get_default_journal_dir()
    if journal_dir is None or not journal_dir.exists():
        raise FileNotFoundError("授業日誌フォルダが見つかりません。--journal-dir で指定してください。")

    schedule_path, target_month = find_schedule_json(repo_dir, args.month)
    events = load_json(schedule_path)

    map_month = repo_dir / f"journal_map_{target_month}.json"
    map_latest = repo_dir / "journal_map_latest.json"
    slots_map = None
    if map_month.exists():
        slots_map = load_json(map_month)
    elif map_latest.exists():
        slots_map = load_json(map_latest)

    slot_map = build_group_slot_map(events, slots_map)

    entries = {}
    with tempfile.TemporaryDirectory() as tmp:
        wb_cache = WorkbookCache(journal_dir, Path(tmp))
        try:
            for ev in events:
                if str(ev.get("date", ""))[:7] != target_month:
                    continue
                key = make_entry_key(ev)
                slot_index = slot_map.get(key, 1)
                entries[key] = extract_entry_for_event(ev, slot_index, wb_cache)
        finally:
            wb_cache.close_all()

    output = {
        "month": target_month,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "entries": entries,
    }

    latest_path = repo_dir / "journal_latest.json"
    month_path = repo_dir / f"journal_{target_month}.json"
    safe_write_json(latest_path, output)
    safe_write_json(month_path, output)

    print(f"[OK] 出力: {latest_path}")
    print(f"[OK] 出力: {month_path}")
    print("[INFO] 抽出対象: content / page / homework / report / recordingUrl / teacher / absence / curriculumProgress / sessionNumber / monthNum / weekNum")


if __name__ == "__main__":
    main()
