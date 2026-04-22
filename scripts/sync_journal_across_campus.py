#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
オンラインペア授業の日誌を本校⇔南教室間でExcelに相互コピーするスクリプト

仕様:
- ONLINE_PAIR_CLASSES に定義された授業だけが対象
- 対面（faceToFace=true）の回はコピーしない
- 片方にだけ内容がある場合にコピー。両方に内容がある場合は上書きしない
- 書き込み前にバックアップを取る
- data_only=False でExcelを開き、数式を保持したまま保存

使い方:
    python sync_journal_across_campus.py                    （自動判定）
    python sync_journal_across_campus.py --month 2026-04    （月を指定）
    python sync_journal_across_campus.py --dry-run           （書き込みせず確認だけ）
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

# ===== オンラインペア対象授業 =====
# (grade, class, subject) の組み合わせ。これ以外はコピーしない。
ONLINE_PAIR_CLASSES: List[Tuple[str, str, str]] = [
    ("e4", "X", "jp"),
    ("e4", "X", "arith"),
    ("j2", "X", "eng"),
    ("j2", "X", "math"),
    ("j3", "X", "eng"),
    ("j3", "X", "math"),
    ("j3", "X", "jp"),
    ("j1", "S", "jp"),
    ("j1", "A", "jp"),
    ("j2", "S", "jp"),
    ("j2", "A", "jp"),
    ("j3", "S", "sci"),
    ("j3", "A", "sci"),
    ("j3", "S", "soc"),
    ("j3", "A", "soc"),
    ("j3", "S", "jp"),
    ("j3", "A", "jp"),
]
ONLINE_PAIR_SET: Set[Tuple[str, str, str]] = set(ONLINE_PAIR_CLASSES)

# ===== 定数（extract_journal_to_json.py と同じ） =====
FIRST_BLOCK_COL = 2
BLOCK_WIDTH = 10
BASE_TOP_MAIN = {"S": 6, "A": 26, "B": 46}
BASE_TOP_X = 6

CAMPUS_JP = {"hon": "本校", "minami": "南教室"}
GRADE_JP = {
    "e4": "小４", "e5": "小５", "e6": "小６",
    "j1": "中１", "j2": "中２", "j3": "中３",
}
SUBJECT_JP = {
    "eng": "英語", "math": "数学", "jp": "国語",
    "sci": "理科", "soc": "社会", "arith": "算数",
}

TARGET_FOLDER_NAME = "09　授業日誌"
BACKUP_DIR_NAME = "_backup"
BACKUP_KEEP_DAYS = 7


# ===== ユーティリティ =====

def get_default_repo_dir() -> Optional[Path]:
    candidates = [
        Path.home() / "OneDrive" / "デスクトップ" / "生徒スケジュール表",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def get_default_journal_dir() -> Optional[Path]:
    candidates = [
        Path.home() / "OneDrive" / "●勉強クラブ共有" / TARGET_FOLDER_NAME,
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def subject_for_filename(grade_key: str, subj_key: str) -> str:
    if subj_key == "math" and grade_key in ("e4", "e5", "e6"):
        return "算数"
    return SUBJECT_JP.get(subj_key, "")


def workbook_filename(campus_code: str, grade_code: str, klass: str, subject_code: str, year: int) -> Optional[str]:
    campus = CAMPUS_JP.get(campus_code)
    grade = GRADE_JP.get(grade_code)
    subject = subject_for_filename(grade_code, subject_code)
    if not campus or not grade or not subject:
        return None
    if klass == "X":
        return f"{campus}{grade}X{subject}_{year}.xlsx"
    return f"{campus}{grade}{subject}_{year}.xlsx"


def find_workbook_path(journal_dir: Path, filename: str) -> Optional[Path]:
    for candidate in journal_dir.rglob(filename):
        if candidate.is_file() and "~$" not in candidate.name and BACKUP_DIR_NAME not in str(candidate):
            return candidate
    return None


def find_month_sheet(wb: openpyxl.Workbook, year: int, month: int):
    target = f"{year:04d}-{month:02d}"
    if target in wb.sheetnames:
        return wb[target]
    # 再試行シート（例: 2026-04(再1)）
    month_re = re.compile(r"^(\d{4})-(\d{2})(?:\(再(\d+)\))?$")
    cands = []
    for name in wb.sheetnames:
        m = month_re.match(name)
        if m and int(m.group(1)) == year and int(m.group(2)) == month:
            cands.append((int(m.group(3) or 0), name))
    if cands:
        cands.sort()
        return wb[cands[-1][1]]
    return None


def merged_top_left_cell(ws, row: int, col: int):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col)
    return ws.cell(row, col)


def read_cell_text(ws, row: int, col: int) -> str:
    cell = merged_top_left_cell(ws, row, col)
    v = cell.value
    return "" if v is None else str(v).strip()


def write_cell(ws, row: int, col: int, value: str):
    """結合セルの左上に書き込む"""
    cell = merged_top_left_cell(ws, row, col)
    cell.value = value


# ===== ブロック読み書き =====

# 各フィールドの (top_row からのオフセット行, left_col からのオフセット列)
FIELD_OFFSETS = {
    "content":      (1, 2),   # top+1, left+2
    "page":         (2, 4),   # top+2, left+4
    "homework1":    (3, 3),   # top+3, left+3
    "homework2":    (4, 3),   # top+4, left+3
    "report":       (5, 3),   # top+5, left+3
    "recordingUrl": (11, 2),  # top+11, left+2
}


def read_block(ws, top_row: int, left_col: int) -> dict:
    content = read_cell_text(ws, top_row + 1, left_col + 2)
    page = read_cell_text(ws, top_row + 2, left_col + 4)
    hw1 = read_cell_text(ws, top_row + 3, left_col + 3)
    hw2 = read_cell_text(ws, top_row + 4, left_col + 3)
    report = read_cell_text(ws, top_row + 5, left_col + 3)
    recording_url = read_cell_text(ws, top_row + 11, left_col + 2)
    return {
        "content": content,
        "page": page,
        "homework1": hw1,
        "homework2": hw2,
        "report": report,
        "recordingUrl": recording_url,
    }


def block_has_content(block: dict) -> bool:
    return any([block["content"], block["page"],
                block["homework1"], block["homework2"],
                block["report"], block["recordingUrl"]])


def write_block(ws, top_row: int, left_col: int, block: dict):
    for field_name, (row_off, col_off) in FIELD_OFFSETS.items():
        value = block.get(field_name, "")
        if value:
            write_cell(ws, top_row + row_off, left_col + col_off, value)


def get_top_row(klass: str) -> Optional[int]:
    if klass == "X":
        return BASE_TOP_X
    return BASE_TOP_MAIN.get(klass)


# ===== スケジュールからペア情報を構築 =====

def load_schedule(repo_dir: Path, target_month: str) -> list:
    month_path = repo_dir / f"schedule_{target_month}.json"
    if month_path.exists():
        return json.loads(month_path.read_text(encoding="utf-8"))
    latest = repo_dir / "schedule_latest.json"
    if latest.exists():
        return json.loads(latest.read_text(encoding="utf-8"))
    raise FileNotFoundError(f"スケジュールJSONが見つかりません: {repo_dir}")


def determine_month_from_schedule(repo_dir: Path) -> str:
    latest = repo_dir / "schedule_latest.json"
    if not latest.exists():
        raise FileNotFoundError(f"schedule_latest.json が見つかりません: {repo_dir}")
    events = json.loads(latest.read_text(encoding="utf-8"))
    months = sorted({str(ev.get("date", ""))[:7] for ev in events if str(ev.get("date", ""))[:7]})
    if not months:
        raise ValueError("年月を判定できません")
    return months[-1]


def build_online_pairs(events: list, target_month: str) -> List[dict]:
    """
    同じ (date, time, grade, class, subject) で本校・南教室の両方にある
    かつ ONLINE_PAIR_SET に含まれる
    かつ両方とも faceToFace=false の授業ペアを返す
    """
    # グループ化: (date, time, grade, class, subject) -> {campus: event}
    groups: Dict[tuple, Dict[str, dict]] = defaultdict(dict)

    for ev in events:
        if str(ev.get("date", ""))[:7] != target_month:
            continue
        key = (ev["date"], ev["time"], ev["grade"], ev["class"], ev["subject"])
        campus = ev.get("campus", "")
        groups[key][campus] = ev

    pairs = []
    for key, campus_map in groups.items():
        grade, klass, subject = key[2], key[3], key[4]
        if (grade, klass, subject) not in ONLINE_PAIR_SET:
            continue
        if "hon" not in campus_map or "minami" not in campus_map:
            continue
        if campus_map["hon"].get("faceToFace") or campus_map["minami"].get("faceToFace"):
            continue
        pairs.append({
            "date": key[0],
            "time": key[1],
            "grade": grade,
            "class": klass,
            "subject": subject,
            "hon_event": campus_map["hon"],
            "minami_event": campus_map["minami"],
        })

    pairs.sort(key=lambda x: (x["grade"], x["class"], x["subject"], x["date"]))
    return pairs


# ===== スロット番号の算出 =====

def build_slot_indices(events: list, target_month: str) -> Dict[str, int]:
    """
    groupKey ごとに、月内での出現順でスロット番号を振る。
    返り値: { "date|time|campus|groupKey|room": slot_index }
    """
    from collections import defaultdict

    def normalize_time(s):
        return str(s or "").replace("~", "～").strip()

    def make_key(ev):
        return "|".join([
            str(ev.get("date", "")),
            normalize_time(ev.get("time", "")),
            str(ev.get("campus", "")),
            str(ev.get("groupKey", "")),
            str(ev.get("room", "")),
        ])

    # groupKey ごとに日付順でソート
    gk_events: Dict[str, list] = defaultdict(list)
    for ev in events:
        if str(ev.get("date", ""))[:7] != target_month:
            continue
        gk = str(ev.get("groupKey", ""))
        gk_events[gk].append(ev)

    result = {}
    for gk, evs in gk_events.items():
        evs.sort(key=lambda x: (x.get("date", ""), x.get("time", "")))
        for i, ev in enumerate(evs):
            result[make_key(ev)] = i + 1

    return result


def event_slot_index(ev: dict, slot_map: Dict[str, int]) -> int:
    def normalize_time(s):
        return str(s or "").replace("~", "～").strip()

    key = "|".join([
        str(ev.get("date", "")),
        normalize_time(ev.get("time", "")),
        str(ev.get("campus", "")),
        str(ev.get("groupKey", "")),
        str(ev.get("room", "")),
    ])
    return slot_map.get(key, 0)


# ===== バックアップ =====

def backup_file(file_path: Path, backup_base: Path, timestamp: str):
    backup_dir = backup_base / BACKUP_DIR_NAME / timestamp
    backup_dir.mkdir(parents=True, exist_ok=True)
    dst = backup_dir / file_path.name
    if not dst.exists():
        shutil.copy2(file_path, dst)
        print(f"  [backup] {file_path.name} → {dst.relative_to(backup_base)}")


def cleanup_old_backups(backup_base: Path, keep_days: int = BACKUP_KEEP_DAYS):
    backup_dir = backup_base / BACKUP_DIR_NAME
    if not backup_dir.exists():
        return
    cutoff = datetime.now().strftime("%Y%m%d")
    # 日付フォルダ名でソートして古いものを削除
    dirs = sorted([d for d in backup_dir.iterdir() if d.is_dir()])
    # keep_days 分の日付数を残す
    date_set = set()
    for d in dirs:
        date_set.add(d.name[:8])
    dates_sorted = sorted(date_set, reverse=True)
    dates_to_remove = dates_sorted[keep_days:]
    for d in dirs:
        if d.name[:8] in dates_to_remove:
            shutil.rmtree(d, ignore_errors=True)
            print(f"  [cleanup] 古いバックアップを削除: {d.name}")


# ===== Workbook キャッシュ（書き込み用: data_only=False） =====

class WritableWorkbookCache:
    def __init__(self, journal_dir: Path):
        self.journal_dir = journal_dir
        self._cache: Dict[str, Tuple[Path, openpyxl.Workbook]] = {}
        self._modified: Set[str] = set()

    def get(self, filename: str) -> Optional[Tuple[Path, openpyxl.Workbook]]:
        if filename in self._cache:
            return self._cache[filename]

        path = find_workbook_path(self.journal_dir, filename)
        if path is None:
            self._cache[filename] = None
            return None

        try:
            wb = openpyxl.load_workbook(path, data_only=False, keep_vba=False)
            self._cache[filename] = (path, wb)
            return (path, wb)
        except Exception as e:
            print(f"  [WARN] {filename} を開けません: {e}")
            self._cache[filename] = None
            return None

    def mark_modified(self, filename: str):
        self._modified.add(filename)

    def save_all(self, backup_base: Path, timestamp: str, dry_run: bool = False):
        import tempfile
        for filename in self._modified:
            entry = self._cache.get(filename)
            if entry is None:
                continue
            path, wb = entry
            if dry_run:
                print(f"  [dry-run] 保存スキップ: {filename}")
                continue
            backup_file(path, backup_base, timestamp)
            # OneDriveのロック回避: 一時ファイルに保存してからコピー
            tmp = Path(tempfile.mktemp(suffix=".xlsx", dir=str(path.parent)))
            try:
                wb.save(tmp)
                wb.close()
                self._cache[filename] = None  # 閉じたのでキャッシュ無効化
                shutil.copy2(tmp, path)
                print(f"  [save] {filename}")
            finally:
                if tmp.exists():
                    tmp.unlink()

    def close_all(self):
        for entry in self._cache.values():
            if entry is not None:
                _, wb = entry
                try:
                    wb.close()
                except Exception:
                    pass


# ===== メイン処理 =====

def sync_journals(
    repo_dir: Path,
    journal_dir: Path,
    target_month: str,
    dry_run: bool = False,
) -> int:
    year, month = map(int, target_month.split("-"))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    print(f"[sync] 対象月: {target_month}")
    print(f"[sync] 日誌フォルダ: {journal_dir}")
    print(f"[sync] dry_run: {dry_run}")
    print()

    # 1. スケジュール読み込み → ペア構築
    events = load_schedule(repo_dir, target_month)
    pairs = build_online_pairs(events, target_month)
    slot_map = build_slot_indices(events, target_month)

    print(f"[sync] オンラインペア: {len(pairs)}件")
    if not pairs:
        print("[sync] コピー対象なし。")
        return 0

    # 2. Excel読み書き
    wb_cache = WritableWorkbookCache(journal_dir)
    copy_count = 0

    try:
        for pair in pairs:
            grade = pair["grade"]
            klass = pair["class"]
            subject = pair["subject"]
            date = pair["date"]

            top_row = get_top_row(klass)
            if top_row is None:
                continue

            # 本校・南教室それぞれのファイル名とスロット
            hon_filename = workbook_filename("hon", grade, klass, subject, year)
            min_filename = workbook_filename("minami", grade, klass, subject, year)
            if not hon_filename or not min_filename:
                continue

            hon_slot = event_slot_index(pair["hon_event"], slot_map)
            min_slot = event_slot_index(pair["minami_event"], slot_map)
            if hon_slot == 0 or min_slot == 0:
                continue

            hon_entry = wb_cache.get(hon_filename)
            min_entry = wb_cache.get(min_filename)
            if hon_entry is None or min_entry is None:
                continue

            hon_path, hon_wb = hon_entry
            min_path, min_wb = min_entry

            hon_ws = find_month_sheet(hon_wb, year, month)
            min_ws = find_month_sheet(min_wb, year, month)
            if hon_ws is None or min_ws is None:
                continue

            hon_left = FIRST_BLOCK_COL + (hon_slot - 1) * BLOCK_WIDTH
            min_left = FIRST_BLOCK_COL + (min_slot - 1) * BLOCK_WIDTH

            hon_block = read_block(hon_ws, top_row, hon_left)
            min_block = read_block(min_ws, top_row, min_left)

            hon_has = block_has_content(hon_block)
            min_has = block_has_content(min_block)

            label = f"{GRADE_JP.get(grade,'')}{klass} {SUBJECT_JP.get(subject,'')} {date}"

            if hon_has and not min_has:
                print(f"  [copy] {label}: 本校 → 南教室")
                if not dry_run:
                    write_block(min_ws, top_row, min_left, hon_block)
                    wb_cache.mark_modified(min_filename)
                copy_count += 1

            elif min_has and not hon_has:
                print(f"  [copy] {label}: 南教室 → 本校")
                if not dry_run:
                    write_block(hon_ws, top_row, hon_left, min_block)
                    wb_cache.mark_modified(hon_filename)
                copy_count += 1

            elif hon_has and min_has:
                # 両方記入済み → 上書きしない
                pass

        # 3. 保存（バックアップ付き）
        if copy_count > 0:
            wb_cache.save_all(journal_dir, timestamp, dry_run=dry_run)
            if not dry_run:
                cleanup_old_backups(journal_dir)

    finally:
        wb_cache.close_all()

    print()
    print(f"[sync] 完了: {copy_count}件コピー{'（dry-run）' if dry_run else ''}")
    return copy_count


def main():
    ap = argparse.ArgumentParser(description="オンラインペア授業の日誌をキャンパス間でコピー")
    ap.add_argument("--repo-dir", help="student-calendar リポジトリの場所")
    ap.add_argument("--journal-dir", help="授業日誌フォルダ")
    ap.add_argument("--month", help="対象月（YYYY-MM）")
    ap.add_argument("--dry-run", action="store_true", help="書き込みせず確認だけ")
    args = ap.parse_args()

    repo_dir = Path(args.repo_dir).resolve() if args.repo_dir else get_default_repo_dir()
    if repo_dir is None or not repo_dir.exists():
        raise FileNotFoundError("student-calendar リポジトリが見つかりません。--repo-dir で指定してください。")

    journal_dir = Path(args.journal_dir) if args.journal_dir else get_default_journal_dir()
    if journal_dir is None or not journal_dir.exists():
        raise FileNotFoundError("授業日誌フォルダが見つかりません。--journal-dir で指定してください。")

    target_month = args.month
    if not target_month:
        target_month = determine_month_from_schedule(repo_dir)

    sync_journals(repo_dir, journal_dir, target_month, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
