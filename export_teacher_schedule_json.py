# -*- coding: utf-8 -*-
"""
先生用スケジュールJSONを書き出すスクリプト

方針:
- 授業の読み取りは、既存の export_by_grade_subject.py の collect_events() をそのまま使う
- 先生判定も、既存の resolve_teacher() / collect_events() をそのまま使う
- ただし、先生凡例の位置だけは月ごとのズレに対応できるよう、
  detect_teacher_legend() を「固定セル」ではなく「行スキャン」に差し替える

出力:
- teacher_schedule_YYYY-MM.json
- teacher_schedule_latest.json
- （未判定がある場合のみ）teacher_schedule_unmatched_YYYY-MM.json

使い方:
  python export_teacher_schedule_json.py --schedule "2026年3月スケジュール.xlsm"
  python export_teacher_schedule_json.py
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string

# 同じフォルダに export_by_grade_subject.py がある前提
import export_by_grade_subject as legacy


# ===== 正規化辞書 =====

CAMPUS_CODE = {
    "本校": "hon",
    "南教室": "minami",
}

GRADE_CODE = {
    "1": "j1",  # 中1
    "2": "j2",  # 中2
    "3": "j3",  # 中3
    "4": "e4",  # 小4
    "5": "e5",  # 小5
    "6": "e6",  # 小6
}

ROOM_MAP = {
    "①": "1",
    "②": "2",
    "③": "3",
    "④": "4",
    "⑤": "5",
    "⑥": "6",
    "⑦": "7",
    "⑧": "8",
    "⑨": "9",
}

ZEN2HAN_TRANS = str.maketrans("０１２３４５６７８９", "0123456789")


# ===== 共通ユーティリティ =====

def normalize_room(room_raw: Any) -> str:
    """
    教室表記を数字文字列に正規化する
    例:
      ① -> "1"
      ２ -> "2"
      第４教室 -> "4"
    """
    if room_raw in (None, ""):
        return ""

    s = str(room_raw).strip()

    for k, v in ROOM_MAP.items():
        if k in s:
            return v

    s2 = s.translate(ZEN2HAN_TRANS)
    m = re.search(r"(\d+)", s2)
    if m:
        return m.group(1)

    return ""


def normalize_subject(grade_digit: str, subj_legacy: str) -> str:
    """
    legacyのsubjは主に「英/数/国/理/社」。
    小学生(4,5,6)の「数」は arith、中学生(1,2,3)の「数」は math。
    """
    s = (subj_legacy or "").strip()
    if s == "英":
        return "eng"
    if s == "国":
        return "jp"
    if s == "理":
        return "sci"
    if s == "社":
        return "soc"
    if s == "数":
        if str(grade_digit) in ("4", "5", "6"):
            return "arith"
        return "math"
    return ""


def grade_label_jp(grade_digit: str) -> str:
    return legacy.GRADE_LABEL.get(str(grade_digit), "")


def subject_label_jp(subject_code: str) -> str:
    return {
        "arith": "算数",
        "math": "数学",
        "eng": "英語",
        "jp": "国語",
        "sci": "理科",
        "soc": "社会",
    }.get(subject_code, "")


def build_group_key(campus_code: str, grade_code: str, klass: str, subject_code: str) -> str:
    k = (klass or "").strip()
    return f"{campus_code}_{grade_code}_{k}_{subject_code}"


def build_label(grade_digit: str, klass: str, subject_code: str) -> str:
    g = grade_label_jp(grade_digit)
    s = subject_label_jp(subject_code)
    k = (klass or "").strip()
    return f"{g}{k} {s}".strip()


def ymd(year: int, month: int, day: Any) -> str:
    d = str(day).strip().translate(ZEN2HAN_TRANS)
    if not d.isdigit():
        return ""
    return f"{year}-{month:02d}-{int(d):02d}"


def _split_lines(cell_text: str) -> List[str]:
    if not cell_text:
        return []
    s = str(cell_text).replace("\r\n", "\n").replace("\r", "\n")
    return [ln.strip() for ln in s.split("\n") if ln.strip() != ""]


def is_face_to_face_cell(text: str) -> bool:
    """
    セル内の2行目に「対面」がある授業を対面扱いにする
    """
    lines = _split_lines(text)
    if len(lines) < 2:
        return False
    return "対面" in lines[1]


def normalize_first_line_for_display(text: str) -> str:
    """
    1行目を表示用に整形（全角数字→半角、全角S/A/B/X→半角大文字、空白除去）
    例: "３Ｓ理" -> "3S理"
    """
    lines = _split_lines(text)
    if not lines:
        return ""
    s = lines[0]
    s = s.translate(ZEN2HAN_TRANS)
    s = s.replace(" ", "").replace("\u3000", "")
    s = s.replace("Ｓ", "S").replace("Ａ", "A").replace("Ｂ", "B").replace("Ｘ", "X")
    return s


def teacher_sort_key(name: str) -> Tuple[int, str]:
    return (1, name) if name == "未判定" else (0, name)


# ===== スケジュールファイル検出 =====

def parse_year_month_from_filename(path: Path) -> Optional[Tuple[int, int]]:
    name = path.name.translate(ZEN2HAN_TRANS)
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", name)
    if not m:
        return None
    year = int(m.group(1))
    month = int(m.group(2))
    if not (1 <= month <= 12):
        return None
    return year, month


def find_schedule_file(schedule_arg: Optional[str] = None) -> Tuple[int, int, Path]:
    here = Path(__file__).resolve().parent

    if schedule_arg:
        p = Path(schedule_arg)
        if not p.is_absolute():
            p = (here / p).resolve()
        if not p.exists():
            raise FileNotFoundError(f"--schedule で指定されたファイルが見つかりません: {p}")
        if p.suffix.lower() != ".xlsm":
            raise ValueError(f".xlsm を指定してください: {p}")
        ym = parse_year_month_from_filename(p)
        if not ym:
            raise ValueError(
                "ファイル名から年月を判定できません。\n"
                "ファイル名を『YYYY年M月...xlsm』（全角数字でも可）にするか、年/月を含めてください。\n"
                f"指定ファイル: {p.name}"
            )
        return ym[0], ym[1], p

    xlsm_files = sorted(here.glob("*.xlsm"))
    if not xlsm_files:
        raise FileNotFoundError("同フォルダに .xlsm が見つかりません。--schedule で指定してください。")

    ym_files: List[Tuple[int, int, Path]] = []
    unknown_files: List[Path] = []
    for f in xlsm_files:
        ym = parse_year_month_from_filename(f)
        if ym:
            ym_files.append((ym[0], ym[1], f))
        else:
            unknown_files.append(f)

    if len(ym_files) == 1:
        return ym_files[0]

    if len(ym_files) >= 2:
        msg = "年月が判定できる .xlsm が複数あります。--schedule で1つ指定してください。\n候補:\n"
        msg += "\n".join([f"- {p.name}" for (_, _, p) in ym_files])
        raise RuntimeError(msg)

    if len(unknown_files) == 1:
        raise ValueError(
            "xlsm は1つ見つかりましたが、ファイル名から年月を判定できません。\n"
            "ファイル名を『YYYY年M月...xlsm』（全角数字でも可）に変更するか、--schedule で指定してください。\n"
            f"見つかったファイル: {unknown_files[0].name}"
        )

    msg = "年月が判定できない .xlsm が複数あります。--schedule で指定してください。\n候補:\n"
    msg += "\n".join([f"- {p.name}" for p in unknown_files])
    raise RuntimeError(msg)


# ===== legacy の先生凡例検出だけ差し替え =====

def patched_detect_teacher_legend(ws) -> Dict[Tuple[str, Any], str]:
    """
    先生凡例を動的に検出する。
    旧コードは H56:AF56 固定だが、実ファイルでは 59行目などにズレることがあるため、
    54〜62行 × C〜AZ の範囲から「先生名らしい＋塗りあり」の密度が最も高い行を採用する。

    それでも見つからない場合は、シート全体から最も密度が高い行を採用する。
    """
    mapping: Dict[Tuple[str, Any], str] = {}

    C = column_index_from_string("C")
    AZ = column_index_from_string("AZ")

    def row_candidates(rr: int) -> List[Tuple[Tuple[str, Any], int, str]]:
        items: List[Tuple[Tuple[str, Any], int, str]] = []
        for cc in range(C, AZ + 1):
            name = legacy.cell_text(ws.cell(rr, cc))
            if not legacy.looks_like_name(name):
                continue
            keys = legacy._fill_color_keys(ws.cell(rr, cc).fill)
            if not keys:
                continue
            for k in keys:
                items.append((k, cc, name))
        return items

    # 第1候補: 54〜62行を優先
    best_items: List[Tuple[Tuple[str, Any], int, str]] = []
    best_score = -1
    for rr in range(54, min(ws.max_row, 62) + 1):
        items = row_candidates(rr)
        score = len(items)
        if score > best_score:
            best_score = score
            best_items = items

    # 第2候補: 見つからなければ全体をゆるく走査
    if best_score <= 0:
        for rr in range(1, ws.max_row + 1):
            items = row_candidates(rr)
            score = len(items)
            if score > best_score:
                best_score = score
                best_items = items

    for k, _cc, name in best_items:
        mapping.setdefault(k, name)

    return mapping


# ここで差し替える
legacy.detect_teacher_legend = patched_detect_teacher_legend


# ===== JSON生成 =====

def convert_events_to_teacher_schedule_json(
    events: List[legacy.Event],
    campus_label: str,
    year: int,
    month: int,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    campus_code = CAMPUS_CODE.get(campus_label, "")
    if not campus_code:
        raise ValueError(f"campus_label が想定外です: {campus_label}")

    out: List[Dict[str, Any]] = []
    unmatched: List[Dict[str, Any]] = []

    for e in events:
        grade_digit = str(e.grade).strip()
        grade_code = GRADE_CODE.get(grade_digit, "")
        if not grade_code:
            continue

        subject_code = normalize_subject(grade_digit, e.subj)
        if not subject_code:
            continue

        klass = (e.klass or "").strip()
        date_str = ymd(year, month, e.day)
        if not date_str:
            continue

        room = normalize_room(e.classroom)
        label = build_label(grade_digit, klass, subject_code)
        gk = build_group_key(campus_code, grade_code, klass, subject_code)

        raw_text = (e.text or "").strip()
        face = is_face_to_face_cell(raw_text)

        display_title = ""
        if face:
            base = normalize_first_line_for_display(raw_text)
            display_title = f"{base}対面" if base else "対面"

        title = display_title or normalize_first_line_for_display(raw_text)

        teacher = (e.teacher or "").strip()
        if not teacher:
            teacher = "未判定"
            unmatched.append(
                {
                    "date": date_str,
                    "time": (e.time or "").strip(),
                    "campusLabel": campus_label,
                    "room": room,
                    "rawText": raw_text,
                    "label": label,
                }
            )

        teacher_label = f"{teacher}先生" if teacher != "未判定" else "未判定"

        out.append(
            {
                "date": date_str,
                "time": (e.time or "").strip(),
                "grade": grade_code,
                "class": klass,
                "subject": subject_code,
                "campus": campus_code,
                "campusLabel": campus_label,
                "room": room,
                "groupKey": gk,
                "label": label,
                "faceToFace": bool(face),
                "displayTitle": display_title,
                # 先生用
                "teacher": teacher,
                "teacherLabel": teacher_label,
                # HTML互換
                "title": title,
                "rawText": raw_text,
            }
        )

    out.sort(
        key=lambda x: (
            x["date"],
            x.get("time", ""),
            teacher_sort_key(x.get("teacher", "")),
            x.get("campus", ""),
            x.get("room", ""),
            x.get("label", ""),
        )
    )
    return out, unmatched


def export_teacher_schedule_json(schedule_arg: Optional[str] = None) -> Tuple[Path, Path]:
    year, month, sch_path = find_schedule_file(schedule_arg)

    # 先生色（塗り）を読むため、data_only=False で開く
    wb = openpyxl.load_workbook(sch_path, data_only=False, keep_vba=True)
    targets = legacy.choose_target_sheets(wb)
    if not targets:
        raise RuntimeError("対象シート（本校/南教室の教務部用）が見つかりません。")

    all_rows: List[Dict[str, Any]] = []
    all_unmatched: List[Dict[str, Any]] = []

    for campus_label, _sheetname, sh in targets:
        events = legacy.collect_events(sh, month)
        rows, unmatched = convert_events_to_teacher_schedule_json(events, campus_label, year, month)
        all_rows.extend(rows)
        all_unmatched.extend(unmatched)

    teachers = sorted({x["teacher"] for x in all_rows}, key=teacher_sort_key)

    payload = {
        "year": year,
        "month": month,
        "teachers": teachers,
        "items": all_rows,
    }

    out_dir = Path(sch_path).parent
    versioned = out_dir / f"teacher_schedule_{year}-{month:02d}.json"
    latest = out_dir / "teacher_schedule_latest.json"

    text = json.dumps(payload, ensure_ascii=False, indent=2)
    versioned.write_text(text, encoding="utf-8")
    latest.write_text(text, encoding="utf-8")

    if all_unmatched:
        warn_path = out_dir / f"teacher_schedule_unmatched_{year}-{month:02d}.json"
        warn_path.write_text(json.dumps(all_unmatched, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[WARN] 担当未判定が {len(all_unmatched)} 件あります: {warn_path}")
    else:
        print("[OK] 担当未判定はありません。")

    return versioned, latest


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--schedule", type=str, default=None, help="対象のスケジュールxlsmを指定（推奨）")
    args = ap.parse_args()

    versioned, latest = export_teacher_schedule_json(args.schedule)
    print(f"[OK] JSONを書き出しました: {versioned}")
    print(f"[OK] JSONを書き出しました: {latest}")


if __name__ == "__main__":
    main()